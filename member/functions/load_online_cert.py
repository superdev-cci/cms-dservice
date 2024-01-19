import datetime
from django.db.models import Sum, Avg
from openpyxl import Workbook
from openpyxl import load_workbook
from member.models import Member
from trip.models import Trip, TripApplication


def main():
    wb = load_workbook('/Users/saintent/Desktop/OnlineCert_19_05_14.xlsx')
    ws = wb.active
    member_pool = []
    for i in range(ws.max_row):
        member = ws['A{}'.format(i + 1)].value
        member_pool.append(member)
        print(member.strip())

    Member.objects.filter(mcode__in=member_pool).update(ocert=1)
    return

def main2():
    wb = load_workbook('/Users/saintent/Desktop/Test.xlsx')
    ws = wb['Sheet1']
    s_pool = []
    d_pool = []
    for i in range(ws.max_row):
        member = ws['A{}'.format(i + 1)].value
        s_pool.append(member)
        # print(member)

    ws = wb['Sheet2']
    for i in range(ws.max_row):
        member = ws['C{}'.format(i + 1)].value
        d_pool.append(member)
    # Member.objects.filter(mcode__in=member_pool).update(ocert=1)
    diff = list(set(d_pool) - set(s_pool))

    for x in diff:
        print(x)

    print(diff)
    return