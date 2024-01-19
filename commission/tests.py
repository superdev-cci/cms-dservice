from django.test import TestCase

# Create your tests here.
from commission.functions.analyze import TransferPvPeriodTree
from commission.functions.analyze.weak_team_commission import WeakTeamTree
from commission.report.pv_expired import PvExpiredReport
from commission.report.pv_activity import PvActivityReport as MemberPvActivityReport
from commission.report.holdpv_activity import PvActivityReport, PvTransferOutReport
from openpyxl import Workbook, styles, load_workbook
from commission.functions import create_status
from commission.report.summary.pv_vs_sale import RoundSummarySalesVsPvReportJSON
from commission.report.weak_strong import WeakStrongBalanceExcelReport
from commission.models import WeakStrongSummary
from commission.report.pv_activity.pv_in_out_with_sponsor_structure_excel import PvActivityWithStructureExcelReport


def test_tree():
    member = 'TH5085119'
    start = '2019-01-01'
    end = '2019-01-07'

    instance = TransferPvPeriodTree(member=member, start=start, end=end)
    data = instance.process()
    print(data)


def pv_expired():
    pv = PvExpiredReport(start='2018-12-01', end='2019-01-31', get_type='daily')
    x = pv.total
    wb = Workbook()
    ws = wb.active
    current_row = 1
    for k, v in x.items():
        print(k, v)
        ws.cell(column=1, row=current_row, value=k)
        pv = v.get('MB', 0)
        ws.cell(column=2, row=current_row, value=pv)
        pv = v.get('FR', 0)
        ws.cell(column=3, row=current_row, value=pv)
        pv = v.get('AG', 0)
        ws.cell(column=4, row=current_row, value=pv)
        current_row += 1

    wb.save('pv_expired.xlsx')
    return


def pv_expired_by_position():
    pv = PvExpiredReport(start='2018-12-01', end='2019-01-31', get_type='daily', group='member__level')
    x = pv.total
    wb = Workbook()
    ws = wb.active
    current_row = 1
    for k, v in x.items():
        print(k, v)
        ws.cell(column=1, row=current_row, value=k)
        pv = v.get('DIS', 0)
        ws.cell(column=2, row=current_row, value=pv)
        pv = v.get('PRO', 0)
        ws.cell(column=3, row=current_row, value=pv)
        pv = v.get('VIP', 0)
        ws.cell(column=4, row=current_row, value=pv)
        current_row += 1

    wb.save('pv_expired_level.xlsx')


def pv_in_out():
    pv = PvActivityReport(start='2018-12-01', end='2019-01-31', get_type='daily')
    x = pv.result
    wb = Workbook()
    ws = wb.active
    current_row = 1
    for k, v in x.items():
        print(k, v)
        ws.cell(column=1, row=current_row, value=k)
        ws.cell(column=2, row=current_row, value=v['in'])
        ws.cell(column=3, row=current_row, value=v['out'])
        ws.cell(column=4, row=current_row, value=v['transfer'])
        current_row += 1

    wb.save('pv_in_out.xlsx')


def pv_group():
    x = PvTransferOutReport(start='2019-01-08', end='2019-01-14', get_type='monthly', mem_code='TH7896456',
                            get_org=True)
    result = x.out_total
    wb = Workbook()
    ws = wb.active
    current_row = 1
    for k, d in result.items():
        print(k, d)
        for target, v in d.items():
            # ws.cell(column=1, row=current_row, value=target)
            print(target, v)
            ws.cell(column=1, row=current_row, value=target)
            current_row += 1
            for s, pv in v.items():
                ws.cell(column=2, row=current_row, value=s)
                ws.cell(column=3, row=current_row, value=pv)
                current_row += 1

    wb.save('pack1.xlsx')


def test_ws_tree():
    start = '2019-01-08'
    end = '2019-01-14'
    member = 'TH0000001'
    depth = 2
    instance = WeakTeamTree(member=member, start=start, end=end)
    data = instance.process(int(depth))

    return data


def test_self_qualify():
    start = '2019-02-01'
    end = '2019-02-28'
    # create_status.update_code(start, end)
    create_status.verify(start, end)


def test_member_activity_report():
    x = ['LA1388191', 'LA1727252', 'LA8525659', 'LA3804841', 'LA2661556', 'LA3979986', 'TH4454551', 'LA7987708',
         'LA4335305', 'LA9909123', 'LA6885585']

    report = MemberPvActivityReport(group=x, get_type='monthly', start='2019-01-01', end='2019-04-30')
    report.process()
    report.save_file()


def create_pv_activity_with_sponsor_structure():
    start_date = '2021-01-01'
    end_date = '2021-02-28'
    x = PvActivityWithStructureExcelReport(start=start_date, end=end_date, group='CE')
    x.process()
    x.save_file()

    x = PvActivityWithStructureExcelReport(start=start_date, end=end_date, group='DE')
    x.process()
    x.save_file()

    x = PvActivityWithStructureExcelReport(start=start_date, end=end_date, group='EE')
    x.process()
    x.save_file()

    x = PvActivityWithStructureExcelReport(start=start_date, end=end_date, group='SE')
    x.process()
    x.save_file()

    x = PvActivityWithStructureExcelReport(start=start_date, end=end_date, group='PE')
    x.process()
    x.save_file()

    x = PvActivityWithStructureExcelReport(start=start_date, end=end_date, group='GL')
    x.process()
    x.save_file()

    x = PvActivityWithStructureExcelReport(start=start_date, end=end_date, group='SI')
    x.process()
    x.save_file()


def create_balance_report():
    start_date = '2021-01-01'
    end_date = '2021-02-28'
    x = WeakStrongBalanceExcelReport(start=start_date, end=end_date, member_group='CE')
    x.process()
    x.save_file()

    x = WeakStrongBalanceExcelReport(start=start_date, end=end_date, member_group='DE')
    x.process()
    x.save_file()

    x = WeakStrongBalanceExcelReport(start=start_date, end=end_date, member_group='EE')
    x.process()
    x.save_file()

    x = WeakStrongBalanceExcelReport(start=start_date, end=end_date, member_group='SE')
    x.process()
    x.save_file()

    x = WeakStrongBalanceExcelReport(start=start_date, end=end_date, member_group='PE')
    x.process()
    x.save_file()


def sales_vs_pv():
    report = RoundSummarySalesVsPvReportJSON(start='2018-01-01', end='2019-12-31')
    wb = Workbook()
    ws = wb.active
    start = 1
    count = 1
    for key, value in report.total.items():
        ws.cell(column=1, row=start, value=count)
        ws.cell(column=2, row=start, value=key)
        ws.cell(column=3, row=start, value=value['pv'])
        ws.cell(column=4, row=start, value=value['sale'])
        start += 1
        count += 1

    wb.save('sales_vs_pv.xlsx')


def check_ws_calculate():
    wb = load_workbook('/Users/saintent/Desktop/WS_Test_result.xlsx')
    ws = wb.active
    fail_count = []
    sum_save = 0
    for i in range(ws.max_row - 1):
        member = ws['A{}'.format(i + 2)].value
        bonus = ws['I{}'.format(i + 2)].value
        ws_data = WeakStrongSummary.objects.filter(mcode=member, rcode=152).first()
        ws_bonus = float(ws_data.total)
        if bonus != ws_bonus:
            fail_count.append('{} -> : CAL {} : Check {}'.format(member, bonus, ws_data.total))
            sum_save += ws_bonus - bonus

    print('Fail test')
    for x in fail_count:
        print(x)

    print(sum_save)
    return

# def result():
# from commission.report.pv_transfer_analyst.pv_transfer_analyst import SummaryPvTransferReceiverAnalyst, SummaryPvTransferAgencyAnalyst
# from ecommerce.models import SaleInvoice
# from django.db.models import Sum, Count, Max, Min, Q
# from django.db.models.functions import TruncMonth, TruncYear, TruncQuarter
#     receiver_view = SummaryPvTransferReceiverAnalyst(start='2019-05-01', end='2019-11-30')
#     receiver_dict = receiver_view.total
#     data1 = dict(filter(lambda elem: elem[1]['total'] == 1, receiver_dict.items()))
#     data2 = dict(filter(lambda elem: elem[1]['total'] == 2, receiver_dict.items()))
#     data3 = dict(filter(lambda elem: elem[1]['total'] == 3, receiver_dict.items()))
#     data4 = dict(filter(lambda elem: elem[1]['total'] == 4, receiver_dict.items()))
#     data5 = dict(filter(lambda elem: elem[1]['total'] == 5, receiver_dict.items()))
#     data6 = dict(filter(lambda elem: elem[1]['total'] == 6, receiver_dict.items()))
#     data7 = dict(filter(lambda elem: elem[1]['total'] == 7, receiver_dict.items()))
#     data = dict(filter(lambda elem: elem[1]['total'] >= 8, receiver_dict.items()))
#
#     # mcode_list1 = set(data1.keys())
#     # mcode_list2 = set(data2.keys())
#     # mcode_list3 = set(data3.keys())
#     # mcode_list4 = set(data4.keys())
#     # mcode_list5 = set(data5.keys())
#     # mcode_list6 = set(data6.keys())
#     # mcode_list7 = set(data7.keys())
#     mcode_list = list(data.keys())
#     mcode_list1 = list(data1.keys())
#     mcode_list2 = list(data2.keys())
#     mcode_list3 = list(data3.keys())
#     mcode_list4 = list(data4.keys())
#     mcode_list5 = list(data5.keys())
#     mcode_list6 = list(data6.keys())
#     mcode_list7 = list(data7.keys())
#
#     mb_si = SaleInvoice.objects.filter(
#         sadate__range=('2019-05-01', '2019-11-30'),
#         mcode__in=mcode_list,
#         cancel=0,
#         sa_type__in=["H", "A"],
#         member__mtype1__in=[0, 1]
#     ).filter(~Q(total=10)).annotate(time=TruncMonth('sadate')).values("mcode", "time").annotate(
#         total_sum_pv=Sum('tot_pv'),
#         total_sum=Sum('total'),
#         max_p=Max('total'),
#         min_p=Min('total'),
#         count=Count('id')
#     ).order_by('mcode', 'time', '-total_sum_pv')
#
#     member_result_set = set([x['mcode'] for x in result])
#     return
# from commission.models import PvTransfer
# summary = PvTransfer.objects.filter(
#     uid__in=receiver_view.agency_list,
#     sadate__range=('2019-05-01', '2019-11-30'),
#     mcode__in=mcode_list8,
#     cancel=0,
#     sa_type__in=["Y",]
# ).aggregate(sum_pv=Sum('tot_pv'))
#
# pool = {}
# for k,v1 in v.items():
#     if len(v1['a/am']) and v1['total'] > 7:
#         pool[k] = {
#             'total_all': v1['total'],
#             'count': 0,
#             'max': [],
#             'min': [],
#             'sum': 0
#         }
#
#         for x in v1['a/am']:
#             pool[k]['count'] += x['count']
#             pool[k]['sum'] += float(x['count']) * x['avg']
#             pool[k]['min'].append(x['min'])
#             pool[k]['max'].append(x['max'])
#
#         pool[k]['min'] = sorted(pool[k]['min'])
#         pool[k]['max'] = sorted(pool[k]['max'], reverse=True)
#
#         pool[k]['min'] = pool[k]['min'][0]
#         pool[k]['max'] = pool[k]['max'][0]
# data1 = dict(filter(lambda elem: elem[1]['count'] > 2 , pool.items()))
#
# from member.models import Member
# terminate_list1 = Member.objects.filter(mcode__in=mcode_list1, status_terminate=1).count()
# terminate_list2 = Member.objects.filter(mcode__in=mcode_list2, status_terminate=1).count()
# terminate_list3 = Member.objects.filter(mcode__in=mcode_list3, status_terminate=1).count()
# terminate_list4 = Member.objects.filter(mcode__in=mcode_list4, status_terminate=1).count()
# terminate_list5 = Member.objects.filter(mcode__in=mcode_list5, status_terminate=1).count()
# terminate_list6 = Member.objects.filter(mcode__in=mcode_list6, status_terminate=1).count()
# terminate_list7 = Member.objects.filter(mcode__in=mcode_list7, status_terminate=1).count()
#
# from member.models import Member
# from commission.models import PvTransfer
# from django.db.models import Q
#
# # PvTransfer.objects.filter(
# #     uid=agency.mcode,
# #     cancel=0,
# #     sadate__range=('2019-05-01', '2019-11-30'),
# #     member__line_lft__gt=agency.line_lft,
# #     member__line_rgt__lt=agency.line_rgt,).count()
#
# pool = {}
# for agency in Member.objects.filter(Q(mtype1=2), ~Q(mcode='TH0123456')):
#     receiver_set = set(PvTransfer.objects.filter(
#         uid=agency.mcode,
#         cancel=0,
#         sadate__range=('2019-05-01', '2019-11-30')
#     ).values_list('mcode', flat=True))
#     child_set = set(agency.down_lines.values_list('mcode', flat=True))
#     set_intersection = receiver_set.intersection(child_set)
#     receiver_count = len(receiver_set)
#     child_receiver_count = len(set_intersection)
#     outer_count = receiver_count - child_receiver_count
#     pool[agency] = {
#         "name": agency.full_name,
#         "receiver" : receiver_count,
#         "child_receiver": child_receiver_count,
#         "outer": outer_count,
#         "ratio": int(100 * (outer_count / receiver_count)),
#     }
# new_pool = sorted(pool.items(), key=lambda item: item[1]['ratio'], reverse=True)