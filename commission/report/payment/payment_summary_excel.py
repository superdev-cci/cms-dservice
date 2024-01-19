from collections import OrderedDict
from datetime import date

from openpyxl.styles import PatternFill
from openpyxl.styles.numbers import BUILTIN_FORMATS

from core.mixin import MonthMixIn
from core.report.excel import ExcelTemplateReport
from .payment_summary import PaymentReport


class PaymentReportExcel(MonthMixIn, ExcelTemplateReport):
    """
    a class for generate excel object that represent member's commission payment.
    Excel object can save to an excel file or response via http request file.
    This class inherit class `ExcelTemplateReport` and `MonthMixIn`
    also you can use a method in those class or overwrite method

    Attributes:
        start (:obj:`date`): a start date of interested data.
        end (:obj:`date`): a end date of interested data.
        get_type (str): time period option [daily, monthly, quarter, yearly]
        mcode (str): (Optional) member code of interested person
    """
    template_file = './templates/report/commission/payment_summary.xlsx'
    fillRed = PatternFill("solid", fgColor='00f5b0af')
    fillYellow = PatternFill("solid", fgColor='00f4e694')
    fillGreen = PatternFill("solid", fgColor='00b4f198')

    class Meta:
        title = 'Total paid summary'
        sheet_name = 'Summary'
        file_name = 'total_paid_summary'
        head_start_col = 4
        head_start_row = 7
        head = {
            'fields': []
        }
        content_start_col = 1
        content_start_row = 8

    def __init__(self, *args, **kwargs):
        super(PaymentReportExcel, self).__init__(*args, **kwargs)
        self.get_type = kwargs.get('get_type', 'monthly')
        self.instance = PaymentReport(**kwargs)
        return

    def create_head(self):
        """
        a method to process data to create Header of table in excel object
        """
        if self.get_type == 'yearly':
            diff = self._calculate_year_range(self.instance.start, self.instance.end)
        else:
            diff = self._calculate_month_range(self.instance.start, self.instance.end)
        start_row = self.Meta.head_start_row
        start_col = self.Meta.head_start_col
        header = []
        text_format = '%Y-%b'
        if self.get_type == 'yearly':
            text_format = '%Y'

        for x in diff:
            header.append('{}-WEEK'.format(x.strftime(text_format)))
            header.append('{}-MONTH'.format(x.strftime(text_format)))
        header.append('TOTAL-WEEK')
        header.append('TOTAL-MONTH')

        self.work_sheet['C3'] = self.Meta.title
        self.work_sheet['D4'] = date.today().strftime('%d/%b/%Y')

        for x in header:
            self.work_sheet.cell(row=start_row, column=start_col).value = x
            start_col += 1

    def build_row_meta(self, row_index, mcode, data, **kwargs):
        """
        a method for index data with cell row

        :param row_index: (int) index of sheet row

        :param mcode: (str) member code of interested person

        :param data: (:obj:`dictionary`) a content write in row

        :return: (:obj:`dictionary`)
        """
        meta = [
            ('no', {'data': row_index, 'alignment': self.style['align_center']}),
            ('mcode', {'data': mcode, 'alignment': self.style['align_left']}),
            ('name', {'data': data['name'], 'alignment': self.style['align_left']}),
        ]

        if self.get_type == 'yearly':
            diff = self._calculate_year_range(self.instance.start, self.instance.end)
        else:
            diff = self._calculate_month_range(self.instance.start, self.instance.end)
        total_week = 0
        total_month = 0
        for x in diff:
            text_format = '%Y-%b'
            if self.get_type == 'yearly':
                text_format = '%Y'

            month = x.strftime(text_format)
            week_data = data['data']['week']
            month_data = data['data']['month']
            current_week = week_data.get(month, None)
            current_month = month_data.get(month, None)
            if current_week is None:
                meta.append(('{}-1'.format(month), {'data': 0, 'alignment': self.style['align_right']}))
            else:
                total_week += current_week
                meta.append(('{}-1'.format(month), {
                    'data': current_week,
                    'alignment': self.style['align_right'],
                    'number_format': BUILTIN_FORMATS[3]
                }))

            if current_month is None:
                meta.append(('{}-2'.format(month), {'data': 0, 'alignment': self.style['align_right']}))
            else:
                total_month += current_month
                meta.append(('{}-2'.format(month), {
                    'data': current_month,
                    'alignment': self.style['align_right'],
                    'number_format': BUILTIN_FORMATS[3]
                }))

        meta.append(('total-1', {
            'data': total_week,
            'alignment': self.style['align_right'],
            'number_format': BUILTIN_FORMATS[3],
            'fill': self.fillGreen
        }))
        meta.append(('total-2', {
            'data': total_month,
            'alignment': self.style['align_right'],
            'number_format': BUILTIN_FORMATS[3],
            'fill': self.fillGreen
        }))
        return OrderedDict(meta)

    def process(self):
        """
        a method to process data in excel object by call method `create_head` and `fill_row`
        """
        count = 1
        current_row = self.Meta.content_start_row
        self.create_head()
        for k, v in self.instance.total.items():
            self.fill_row(count, k, v, row=current_row)
            count += 1
            current_row += 1

    @property
    def file_name(self):
        """
        a method to named a excel object

        :return: (str): excel object's file name
        """
        name = '{}_{}'.format(
            self.Meta.file_name, date.today().strftime('%d-%b-%Y'))
        return '{}.xlsx'.format(name)
