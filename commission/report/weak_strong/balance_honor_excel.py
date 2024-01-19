from abc import ABC
from collections import OrderedDict
from datetime import date
from openpyxl.styles import PatternFill
from openpyxl.styles.numbers import FORMAT_NUMBER_COMMA_SEPARATED1, BUILTIN_FORMATS
from core.mixin import MonthMixIn
from core.report.excel import ExcelTemplateReport
from commission.report.weak_strong.weak_strong_summary import WeakStrongBalanceJsonReport
from commission.functions.honor import HonorQualify


class WeakStrongBalanceExcelReport(MonthMixIn, ExcelTemplateReport, ABC):
    template_file = './templates/report/commission/pv_export_import_group.xlsx'
    fillRed = PatternFill("solid", fgColor='00f5b0af')
    fillYellow = PatternFill("solid", fgColor='00f4e694')
    fillGreen = PatternFill("solid", fgColor='00b4f198')

    class Meta:
        title = 'Weak team balance'
        sheet_name = 'Summary'
        file_name = 'weak_team_balance'
        head_start_col = 5
        head_start_row = 7
        head = {
            'fields': []
        }
        content_start_col = 1
        content_start_row = 8

    def __init__(self, *args, **kwargs):
        super(WeakStrongBalanceExcelReport, self).__init__(*args, **kwargs)
        self.get_type = kwargs.get('get_type', None)
        self.member_group = kwargs.get('member_group', None)
        self.balance_pool = WeakStrongBalanceJsonReport(*args, **kwargs)
        self.start = self.balance_pool.start
        self.end = self.balance_pool.end

    def create_head(self):
        diff = self.month_diff_range(self.end, self.start)
        fields = []
        for x in diff:
            fields.append('{}-BALANCE'.format(x.strftime('%b')))
            fields.append('{}-QUALIFY'.format(x.strftime('%b')))
            # self.Meta.head['fields'].append('{}-DIFF'.format(x.strftime('%Y-%b')))
        # self.Meta.head['fields'].append('TOTAL-IN')
        # self.Meta.head['fields'].append('TOTAL-OUT')
        # self.Meta.head['fields'].append('TOTAL-DIFF')
        fields.append('TOTAL-BALANCE')
        fields.append('MAX-QUALIFY')

        self.work_sheet['C3'] = self.Meta.title
        self.work_sheet['D4'] = date.today().strftime('%d/%b/%Y')
        super(WeakStrongBalanceExcelReport, self).create_head()
        start_row = self.Meta.head_start_row
        start_col = self.Meta.head_start_col
        for x in fields:
            self.work_sheet.cell(row=start_row, column=start_col).value = x
            self.work_sheet.cell(row=start_row, column=start_col).border = self.style['full_border']
            start_col += 1

    def build_row_meta(self, row_index, mcode, data, **kwargs):
        meta = [
            ('no', {'data': row_index, 'alignment': self.style['align_center']}),
            ('mcode', {'data': mcode, 'alignment': self.style['align_left']}),
            ('name', {'data': data['name'], 'alignment': self.style['align_left']}),
            ('honor', {'data': data['honor'], 'alignment': self.style['align_center']})
        ]
        diff = self.month_diff_range(self.end, self.start)
        total = 0
        qualify_level = ''
        for x in diff:
            month = x.strftime('%Y-%b')
            current_month = data.get(month, None)
            if current_month is None:
                meta.append(('{}-1'.format(month), {'data': 0, 'alignment': self.style['align_right']}))
                meta.append(('{}-2'.format(month), {'data': '-', 'alignment': self.style['align_right']}))
                # meta.append(('{}-3'.format(month), {'data': 0, 'alignment': self.style['align_right']}))
            else:
                balance = current_month.get('balance', 0)
                qualify = current_month.get('qualify', '')
                total = total + balance
                meta.append(('{}-1'.format(month), {
                    'data': balance,
                    'alignment': self.style['align_right'],
                    'number_format': BUILTIN_FORMATS[3]
                }))
                meta.append(('{}-2'.format(month), {
                    'data': qualify,
                    'alignment': self.style['align_right'],
                }))
                qualify_level = HonorQualify.compare_qualify(qualify_level, qualify)

        meta.append(('total-1', {
            'data': total,
            'alignment': self.style['align_right'],
            'number_format': BUILTIN_FORMATS[3],
            'fill': self.fillGreen
        }))
        meta.append(('total-2', {
            'data': qualify_level,
            'alignment': self.style['align_right'],
            'fill': self.fillGreen
        }))

        return OrderedDict(meta)

    def process(self):
        count = 1
        current_row = self.Meta.content_start_row
        self.create_head()

        for k, v in self.balance_pool.total.items():
            self.fill_row(count, k, v, row=current_row)
            count += 1
            current_row += 1
        return

    @property
    def file_name(self):
        name = '{}_{}_{}_{}'.format(
            self.start.strftime('%Y-%b'),
            self.end.strftime('%Y-%b'),
            self.Meta.file_name, self.member_group)
        return '{}.xlsx'.format(name)
