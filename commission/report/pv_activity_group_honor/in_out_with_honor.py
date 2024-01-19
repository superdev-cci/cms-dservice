from abc import ABC
from collections import OrderedDict
from datetime import date

from openpyxl.styles import PatternFill
from openpyxl.styles.numbers import FORMAT_NUMBER_COMMA_SEPARATED1, BUILTIN_FORMATS

from commission.report.pv_activity_group_honor.bought_hold_honor import PvHoldInReport
from commission.report.pv_activity_group_honor.in_honor_group import PvTransferInReport
from commission.report.pv_activity_group_honor.out_honor_group import PvTransferOutReport
from core.mixin import MonthMixIn
from core.report.excel import ExcelTemplateReport


class PvActivityHonorGroupReport(MonthMixIn, ExcelTemplateReport, ABC):
    template_file = './templates/report/commission/pv_export_import_group.xlsx'
    fillRed = PatternFill("solid", fgColor='00f5b0af')
    fillYellow = PatternFill("solid", fgColor='00f4e694')
    fillGreen = PatternFill("solid", fgColor='00b4f198')

    class Meta:
        title = 'PV IN/OUT'
        sheet_name = 'Summary'
        file_name = 'pv_in_out_summary'
        head_start_col = 5
        head_start_row = 7
        head = {
            'fields': []
        }
        content_start_col = 1
        content_start_row = 8

    def __init__(self, *args, **kwargs):
        super(PvActivityHonorGroupReport, self).__init__(*args, **kwargs)
        self.get_type = kwargs.get('get_type', None)
        self.member_group = kwargs.get('group', None)
        self.transfer_in = PvTransferInReport(*args, **kwargs)
        self.transfer_out = PvTransferOutReport(*args, **kwargs)
        self.bought_in = PvHoldInReport(*args, **kwargs)
        self.start = self.transfer_in.start
        self.end = self.transfer_in.end

    def create_head(self):
        diff = self.month_diff_range(self.transfer_in.end, self.transfer_in.start)
        for x in diff:
            self.Meta.head['fields'].append('{}-IN'.format(x.strftime('%Y-%b')))
            self.Meta.head['fields'].append('{}-OUT'.format(x.strftime('%Y-%b')))
            # self.Meta.head['fields'].append('{}-DIFF'.format(x.strftime('%Y-%b')))
        self.Meta.head['fields'].append('TOTAL-IN')
        self.Meta.head['fields'].append('TOTAL-OUT')
        self.Meta.head['fields'].append('TOTAL-DIFF')

        self.work_sheet['C3'] = self.Meta.title
        self.work_sheet['D4'] = date.today().strftime('%d/%b/%Y')
        super(PvActivityHonorGroupReport, self).create_head()

    def build_row_meta(self, row_index, mcode, data, **kwargs):
        meta = [
            ('no', {'data': row_index, 'alignment': self.style['align_center']}),
            ('mcode', {'data': mcode, 'alignment': self.style['align_left']}),
            ('name', {'data': data['name'], 'alignment': self.style['align_left']}),
            ('honor', {'data': data['honor'], 'alignment': self.style['align_center']})
        ]
        diff = self.month_diff_range(self.transfer_in.end, self.transfer_in.start)
        total_in = 0
        total_out = 0
        for x in diff:
            month = x.strftime('%Y-%b')
            current_month = data['data'].get(month, None)
            if current_month is None:
                meta.append(('{}-1'.format(month), {'data': 0, 'alignment': self.style['align_right']}))
                meta.append(('{}-2'.format(month), {'data': 0, 'alignment': self.style['align_right']}))
                # meta.append(('{}-3'.format(month), {'data': 0, 'alignment': self.style['align_right']}))
            else:
                pv_in = current_month.get('pv_in', 0)
                pv_out = current_month.get('pv_out', 0)
                pv_hold = current_month.get('bought', 0)
                total_in += (pv_in + pv_hold)
                total_out += pv_out
                meta.append(('{}-1'.format(month), {
                    'data': pv_in + pv_hold,
                    'alignment': self.style['align_right'],
                    'number_format': BUILTIN_FORMATS[3]
                }))
                meta.append(('{}-2'.format(month), {
                    'data': pv_out,
                    'alignment': self.style['align_right'],
                    'number_format': BUILTIN_FORMATS[3]
                }))

        meta.append(('total-1', {
            'data': total_in,
            'alignment': self.style['align_right'],
            'number_format': BUILTIN_FORMATS[3],
            'fill': self.fillGreen
        }))
        meta.append(('total-2', {
            'data': total_out,
            'alignment': self.style['align_right'],
            'number_format': BUILTIN_FORMATS[3],
            'fill': self.fillYellow
        }))
        meta.append(('total-3', {
            'data': total_in - total_out,
            'alignment': self.style['align_right'],
            'number_format': BUILTIN_FORMATS[3],
            'fill': self.fillRed
        }))

        return OrderedDict(meta)

    def add_to_pool(self, pool, items, key):
        for k, v in items.items():
            if k not in pool:
                pool[k] = {
                    'name': v['name'],
                    'honor': v['honor'],
                    'data': {}
                }

            for k1, v1 in v['data'].items():
                if k1 in pool[k]['data']:
                    pool[k]['data'][k1][key] = v1
                else:
                    pool[k]['data'][k1] = {
                        key: v1
                    }
        return

    def process(self):
        count = 1
        current_row = self.Meta.content_start_row
        self.create_head()
        pool = {}
        self.add_to_pool(pool, self.bought_in.total, 'bought')
        self.add_to_pool(pool, self.transfer_in.total, 'pv_in')
        self.add_to_pool(pool, self.transfer_out.total, 'pv_out')
        for k, v in pool.items():
            self.fill_row(count, k, v, row=current_row)
            count += 1
            current_row += 1
        return

    @property
    def file_name(self):
        name = '{}_{}_{}_{}'.format(
            self.start.strftime('%Y-%b'),
            self.end.strftime('%Y-%b'),
            self.Meta.file_name, self.member_group)
        return '{}.xlsx'.format(name)
