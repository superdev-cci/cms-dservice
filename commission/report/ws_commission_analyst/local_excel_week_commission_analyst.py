from openpyxl import Workbook
from commission.report.ws_commission_analyst.week_commission_analyst import WeekCommissionAnalyst
from datetime import datetime, timedelta
import statistics


def export_ws_bonus(start_date="2019-01-01", end_date="2019-12-31"):
    sd_obj = datetime.strptime(start_date, "%Y-%m-%d").date()
    ed_obj = datetime.strptime(end_date, "%Y-%m-%d").date()
    report = WeekCommissionAnalyst(start=start_date, end=end_date).total
    wb = Workbook()
    sht = wb.active
    sht.title = "ws_bonus"
    # header render
    sht.merge_cells("A1:A2")
    sht["A1"] = "รหัส"
    sht.merge_cells("B1:B2")
    sht["B1"] = "ผู้ทำรายการ"
    sht.merge_cells("C1:C2")
    sht["C1"] = "วันที่เข้าผัง"
    sht.merge_cells("D1:I1")
    sht["D1"] = "All Summary"
    sht["D2"] = "min"
    sht["E2"] = "max"
    sht["F2"] = "avg"
    sht["G2"] = "sum"
    sht["H2"] = "std"
    sht["I2"] = "median"
    head_col_start = 10
    month_list = WeekCommissionAnalyst.month_diff_range(ed_obj, sd_obj)
    for index, mth in enumerate(month_list):
        col_shift = 4 * index
        sht.merge_cells(start_column=(head_col_start + col_shift), end_column=(head_col_start + col_shift + 3),
                        start_row=1, end_row=1)
        sht.cell(column=(head_col_start + col_shift), row=1, value=mth.strftime("%d/%m/%Y"))
        sht.cell(column=(head_col_start + col_shift), row=2, value="min")
        sht.cell(column=(head_col_start + col_shift + 1), row=2, value="max")
        sht.cell(column=(head_col_start + col_shift + 2), row=2, value="avg")
        sht.cell(column=(head_col_start + col_shift + 3), row=2, value="sum")
    # body render data
    row_start = 3
    for mcode, data_dict in report.items():
        sht.cell(column=1, row=row_start, value=mcode)
        sht.cell(column=2, row=row_start, value=data_dict["name"])
        sht.cell(column=3, row=row_start, value=data_dict["distributor_date"])
        min_val = min([x['min'] for x in data_dict['data_record'].values()])
        max_val = max([x['max'] for x in data_dict['data_record'].values()])
        avg_val = statistics.mean([x['avg'] for x in data_dict['data_record'].values()])
        sum_val = sum([x['sum'] for x in data_dict['data_record'].values()])
        avg_list = [x['avg'] for x in data_dict['data_record'].values()]
        if len(avg_list) >= 2:
            std_val = statistics.stdev(avg_list)
        else:
            std_val = 0
        if len(avg_list) >= 1:
            median_val = statistics.median(avg_list)
        else:
            median_val = 0
        sht.cell(column=4, row=row_start, value=min_val)
        sht.cell(column=5, row=row_start, value=max_val)
        sht.cell(column=6, row=row_start, value=avg_val)
        sht.cell(column=7, row=row_start, value=sum_val)
        sht.cell(column=8, row=row_start, value=std_val)
        sht.cell(column=9, row=row_start, value=median_val)
        col_start = 10
        for k_time, bonus in data_dict["data_record"].items():
            col_shift = (4 * month_list.index(k_time))
            sht.cell(column=(col_start + col_shift), row=row_start, value=bonus["min"])
            sht.cell(column=(col_start + col_shift + 1), row=row_start, value=bonus["max"])
            sht.cell(column=(col_start + col_shift + 2), row=row_start, value=bonus["avg"])
            sht.cell(column=(col_start + col_shift + 3), row=row_start, value=bonus["sum"])
        row_start = sht.max_row + 1
    wb.save("/home/jew/Desktop/WS_BonusAnalyst2.xlsx")
