import datetime
from django.db.models import Sum, Avg
from openpyxl import Workbook
from openpyxl import load_workbook
from member.models import Member
from event.models import PreAttendee, Event, Attendee
from trip.models import Trip, TripApplication
from openpyxl import load_workbook
import re

def main():
    # wb = load_workbook('/Users/saintent/Desktop/CCI_ECON.xlsx')
    #     # ws = wb.active
    #     # member_pool = []
    #     # for i in range(ws.max_row):
    #     #     member = ws['H{}'.format(i + 1)].value.strip(' ')
    #     #     member_pool.append(member)
    member_pool = ['TH1929496', 'TH4356994', 'TH8887444', 'TH7706142', 'TH2475721', 'TH8838233', 'TH1964255',
                   'TH8883439', 'TH6059333', 'TH6539589', 'TH4508231',
                   'TH1969251', 'TH7599929', 'TH3398461', 'TH4905256', 'TH6600067', 'TH8933276', 'TH2005844',
                   'TH7671588', 'TH8928302', 'TH6239270', 'TH9895961',
                   'TH1734332', 'TH5246695', 'TH6245915', 'TH7821108', 'TH3752021', 'TH9198682', 'TH9779815',
                   'TH7254203', 'TH3298999', 'TH2222250', 'TH8599965',
                   'TH4586486', 'TH2222214', 'TH4563644', 'TH4485436', 'TH8484469', 'TH5768178', 'TH6677569',
                   'TH9888999', 'TH4056877', 'TH4709472', 'TH2112995',
                   'TH9999529', 'TH9288256', 'TH7223510', 'TH6958618', 'TH0704657', 'TH2056910', 'TH2365439',
                   'TH5168168', 'TH2659394', 'TH6844422', 'TH6600066',
                   'TH2047169', 'TH6999995', 'TH4976122', 'TH8198918', 'TH9889999', 'TH9889899', 'TH5139154',
                   'TH3269172', 'TH5454999', 'TH4609985', 'TH1972999',
                   'TH9903699', 'TH3554893', 'TH2040498', 'TH6339519', 'TH8367981', 'TH3610672', 'TH8646872',
                   'TH3725891', 'TH5664544', 'TH2644656', 'TH0838352',
                   'TH2456491', 'TH9280215', 'TH1689107', 'TH6654249', 'TH2513505', 'TH9998182', 'TH8982299',
                   'TH5250549', 'TH4996887', 'TH8665222', 'TH5454141',
                   'TH6660889', 'TH9051970', 'TH2222222', 'TH8401480', 'TH4833439', 'TH0099395', 'TH9381564',
                   'TH4944966', 'TH7615305', 'TH3509528', 'TH9329621',
                   'TH9420786', 'TH8507160', 'TH8925819', 'TH9407348', 'TH8398997', 'TH9090989', 'TH8759184',
                   'TH7756987', 'TH7978969', 'TH3627989', 'TH1905556',
                   'TH2465454', 'TH1911432', 'TH3651456', 'TH5649929', 'TH2222244', 'TH2459168', 'TH5112304',
                   'TH7389678', 'TH0201888', 'TH6754803', 'TH1111110',
                   'TH9909968', 'TH3296789', 'TH2466495', 'TH4795547', 'TH9648511', 'TH5870049', 'TH2288995',
                   'TH2793858', 'TH3346862', 'TH1740722', 'TH0072436',
                   'TH5722656', 'TH2458168', 'TH6299789', 'TH1270246', 'TH7238634', 'TH8116059', 'TH1478633',
                   'TH2643868', 'TH9898899', 'TH2222236', 'TH4548339',
                   'TH4800610', 'TH1920953', 'TH9064517', 'TH5599904', 'TH9999979', 'TH5858789', 'TH8891020',
                   'TH4456691', 'TH8782947', 'TH3938011', 'TH9317489',
                   'TH1541361', 'TH3449889', 'TH8068533', 'TH2396261', 'TH5433959', 'TH6175494', 'TH1988755',
                   'TH2522456', 'TH2788803', 'TH6493263', 'TH8454995',
                   'TH6577168', 'TH2222247', 'TH2456449', 'TH9169173', 'TH9999308', 'TH2519889', 'TH2425262',
                   'TH0568140', 'TH2047812', 'TH5962174', 'TH8245589',
                   'TH1637849', 'TH4564592', 'TH3910532', 'TH1168477', 'TH3111270', 'TH8935244', 'TH8864440',
                   'TH5423583', 'TH6719511', 'TH0224756', 'TH7354399',
                   'TH2996685', 'TH4545459', 'TH2936828', 'TH2979417', 'TH4101562', 'TH0016699', 'TH4501753',
                   'TH4772033', 'TH4926452', 'TH6666772', 'TH8208492',
                   'TH6600099', 'TH2418837', 'TH8100415', 'TH5592738', 'TH0061670', 'TH6380560', 'TH9714288',
                   'TH4279245', 'TH5945865', 'TH7333323', 'TH0520331',
                   'TH3865009', 'TH5865173', 'TH5350511', 'TH5498401', 'TH2619742', 'TH6956264', 'TH8557125',
                   'TH3000000', 'TH2000000', 'TH6830746', 'TH7896456',
                   'TH2644655', 'TH1265258', 'TH9189239', 'TH2222245', 'TH6615562', 'TH2644885', 'TH4261655',
                   'TH1119989', 'LA1692176', 'TH5981798', 'TH6998565',
                   'TH9063090', 'TH5971069', 'TH1733477', 'TH1699768', 'TH6310472', 'TH5029699', 'TH4102161',
                   'TH8960571', 'TH1000000', 'TH7778999', 'TH4686924',
                   'LA0172862', 'TH4112153', 'TH1100644', 'TH1156678', 'TH3297789', 'LA1804688', 'TH5117981',
                   'TH0707839', 'TH3333339', 'TH1536007', 'TH4623698',
                   'TH2689699', 'TH1792329', 'TH4752525', 'TH2632586', 'TH7118991']
    member_pool = list(set(member_pool))
    print(member_pool)
    queryset = Member.objects.filter(mcode__in=member_pool)
    pre_attendee = PreAttendee.objects.get(event__id=1)
    # print(queryset.query)
    # all = [x.code for x in queryset]
    for x in queryset:
        # print('Test' , x)
        # queryset = Member.objects.get(mcode=x)
        pre_attendee.members.add(x)
        print(x.code)

    return


def import_manual_event():
    wb = load_workbook('/Users/saintent/Desktop/event.xlsx')
    ws = wb.active
    member_pool = []

    event = Event.objects.get(event_tag="TBN01")
    if hasattr(event, 'attendee'):
        attendee = event.attendee
    else:
        attendee = Attendee.objects.create(event=event)

    print(event)

    for i in range(ws.max_row):
        member = ws['B{}'.format(i + 1)].value.upper()
        member_pool.append(member)
        match = re.findall('\d+', member)
        new_mcode = 'TH{}'.format(match[0])
        print(new_mcode, 'enter')
        is_exist = attendee.members.filter(mcode=new_mcode)
        if len(is_exist) is 0:
            try:
                member_instance = Member.objects.get(mcode=new_mcode)
                attendee.members.add(member_instance)
                print(member.strip(), 'Add to event')
            except Exception as e:
                print(member.strip(), 'Add fail')
        else:
            print(member.strip(), 'Has register')
    return
