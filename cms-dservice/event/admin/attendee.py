from django.conf.urls import url
from django.contrib import admin, messages
from django.http import HttpResponse
from django.shortcuts import redirect
from django.urls import reverse
from django.utils.html import format_html
from openpyxl import Workbook, styles
from openpyxl.writer.excel import save_virtual_workbook

from member.models import Member
from ..models import Attendee, PreAttendee


class MemberInline(admin.TabularInline):
    model = PreAttendee.members.through
    # readonly_fields = ['PreAttendee_members']
    # raw_id_fields = ('event_pre_attendee',)
    extra = 1


@admin.register(Attendee)
class AttendeeAdmin(admin.ModelAdmin):
    list_display = ('event',)
    fields = ('event',)
    raw_id_fields = ('event',)
    # exclude = ('memnber',)
    # inlines = [MemberInline]


@admin.register(PreAttendee)
class PreAttendeeAdmin(admin.ModelAdmin):
    list_display = ('event', 'group', 'event_actions')
    fields = ('event',)
    raw_id_fields = ('event',)
    # inlines = [MemberInline]

    def get_urls(self):
        urls = super().get_urls()
        custom_urls = [
            url(
                r'^(?P<pk>.+)/report/$',
                self.admin_site.admin_view(self.process_preadttendee_action),
                name='event-preattendee',
            ),
        ]
        return custom_urls + urls

    def event_actions(self, obj):
        return format_html(
            '<a class="button" href="{}">Report</a>',
            reverse('admin:event-preattendee', args=[obj.pk]),
        )

    def process_preadttendee_action(self, request, pk, *args, **kwargs):
        try:
            obj = PreAttendee.objects.get(id=pk)
            wb = Workbook()
            ws = wb.active
            ws['A1'] = obj.event.name
            count = 1
            for x in obj.members.all():
                ws.cell(column=1, row=count + 1, value=count)
                ws.cell(column=2, row=count + 1, value=x.code)
                ws.cell(column=3, row=count + 1, value=x.full_name)
                ws.cell(column=4, row=count + 1, value=x.get_level())
                count += 1
            response = HttpResponse(save_virtual_workbook(wb),
                                    content_type='application/ms-excel')
            response['Content-Disposition'] = 'attachment; filename=event_{}.xlsx'.format(obj.event.name)
            self.message_user(request, 'Complete', level=messages.SUCCESS)
            return response
        except Exception as e:
            pass
        return redirect('/admin/event/preattendee/')
