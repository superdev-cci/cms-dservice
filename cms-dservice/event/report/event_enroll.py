from openpyxl import Workbook, styles
from event.models import Event


def main():
    start = 2
    wb = Workbook()
    ws = wb.active
    event = Event.objects.first()
    for x in event.attendee.members.all():
        print(x.mcode, x.name_t, x.honor)
        ws.cell(column=1, row=start, value=x.mcode)
        ws.cell(column=2, row=start, value=x.full_name)
        ws.cell(column=3, row=start, value=x.level)
        ws.cell(column=4, row=start, value=x.honor)
        start += 1

    wb.save('event.xlsx')
