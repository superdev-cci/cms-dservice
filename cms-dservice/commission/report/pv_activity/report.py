from openpyxl.styles import PatternFill

from commission.report.pv_activity.pv_bought import PvHoldInReport
from commission.report.pv_activity.pv_in import PvTransferInReport
from commission.report.pv_activity.pv_out import PvTransferOutReport
from commission.report.pv_activity_group_honor import PvActivityHonorGroupReport


class PvActivityReport(PvActivityHonorGroupReport):
    template_file = './templates/report/commission/pv_export_import_group.xlsx'
    fillRed = PatternFill("solid", fgColor='00f5b0af')
    fillYellow = PatternFill("solid", fgColor='00f4e694')
    fillGreen = PatternFill("solid", fgColor='00b4f198')

    class Meta:
        title = 'PV IN/OUT'
        sheet_name = 'Summary'
        file_name = 'pv_in_out_summary'
        head_start_col = 5
        head_start_row = 7
        head = {
            'fields': []
        }
        content_start_col = 1
        content_start_row = 8

    def __init__(self, *args, **kwargs):
        super(PvActivityReport, self).__init__(*args, **kwargs)
        self.member_group = kwargs.get('group', None)
        self.transfer_in = PvTransferInReport(*args, **kwargs)
        self.transfer_out = PvTransferOutReport(*args, **kwargs)
        self.bought_in = PvHoldInReport(*args, **kwargs)
        self.start = self.transfer_in.start
        self.end = self.transfer_in.end
