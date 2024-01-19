from abc import ABC
from collections import OrderedDict
from datetime import date

from openpyxl.styles import PatternFill
from openpyxl.styles.numbers import FORMAT_NUMBER_COMMA_SEPARATED1, BUILTIN_FORMATS
from core.mixin import MonthMixIn
from core.report.excel import ExcelTemplateReport
from commission.report.pv_activity.pv_in_out_with_sponsor_structure import PvInOutWithSponsorStructureJSON


class PvActivityWithStructureExcelReport(MonthMixIn, ExcelTemplateReport, ABC):
    template_file = './templates/report/commission/pv_export_import_group.xlsx'
    fillRed = PatternFill("solid", fgColor='00f5b0af')
    fillYellow = PatternFill("solid", fgColor='00f4e694')
    fillGreen = PatternFill("solid", fgColor='00b4f198')

    class Meta:
        title = 'PV IN/OUT WITH SPONSOR STRUCTURE'
        sheet_name = 'Summary'
        file_name = 'pv_in_out_with_sponsor_structure_summary'
        head_start_col = 5
        head_start_row = 7
        head = {
            'fields': []
        }
        content_start_col = 1
        content_start_row = 8

    def __init__(self, *args, **kwargs):
        super(PvActivityWithStructureExcelReport, self).__init__(*args, **kwargs)
        self.get_type = kwargs.get('get_type', None)
        self.member_group = kwargs.get('group', None)
        self.data = PvInOutWithSponsorStructureJSON(**kwargs)

    def create_head(self):
        diff = self.month_diff_range(self.data.end, self.data.start)
        start_row = self.Meta.head_start_row
        start_col = self.Meta.head_start_col
        header = []
        for x in diff:
            header.append('{}-SALES'.format(x.strftime('%Y-%b')))
            header.append('{}-IN'.format(x.strftime('%Y-%b')))
            header.append('{}-OUT'.format(x.strftime('%Y-%b')))
        header.append('TOTAL-SALES')
        header.append('TOTAL-IN')
        header.append('TOTAL-OUT')

        self.work_sheet['C3'] = self.Meta.title
        self.work_sheet['D4'] = date.today().strftime('%d/%b/%Y')

        for x in header:
            self.work_sheet.cell(row=start_row, column=start_col).value = x
            start_col += 1
        return

    def build_row_meta(self, row_index, mcode, data, **kwargs):
        meta = [
            ('no', {'data': row_index, 'alignment': self.style['align_center']}),
            ('mcode', {'data': mcode, 'alignment': self.style['align_left']}),
            ('name', {'data': data['name'], 'alignment': self.style['align_left']}),
            ('honor', {'data': data['honor'], 'alignment': self.style['align_center']})
        ]
        diff = self.month_diff_range(self.data.end, self.data.start)
        total_sales = 0
        total_in = 0
        total_out = 0
        for x in diff:
            month = x.strftime('%Y-%b')
            current_month = data['data'].get(month, None)
            if current_month is None:
                meta.append(('{}-1'.format(month), {'data': 0, 'alignment': self.style['align_right']}))
                meta.append(('{}-2'.format(month), {'data': 0, 'alignment': self.style['align_right']}))
                meta.append(('{}-3'.format(month), {'data': 0, 'alignment': self.style['align_right']}))
            else:
                sales = current_month.get('sales', 0)
                pv_in = current_month.get('in', 0)
                pv_out = current_month.get('out', 0)
                total_in += pv_in
                total_out += pv_out
                total_sales += sales
                meta.append(('{}-1'.format(month), {
                    'data': sales,
                    'alignment': self.style['align_right'],
                    'number_format': BUILTIN_FORMATS[3]
                }))
                meta.append(('{}-2'.format(month), {
                    'data': pv_in,
                    'alignment': self.style['align_right'],
                    'number_format': BUILTIN_FORMATS[3]
                }))
                meta.append(('{}-3'.format(month), {
                    'data': pv_out,
                    'alignment': self.style['align_right'],
                    'number_format': BUILTIN_FORMATS[3]
                }))
        meta.append(('total-1', {
            'data': total_sales,
            'alignment': self.style['align_right'],
            'number_format': BUILTIN_FORMATS[3],
            'fill': self.fillGreen
        }))
        meta.append(('total-2', {
            'data': total_in,
            'alignment': self.style['align_right'],
            'number_format': BUILTIN_FORMATS[3],
            'fill': self.fillGreen
        }))
        meta.append(('total-3', {
            'data': total_out,
            'alignment': self.style['align_right'],
            'number_format': BUILTIN_FORMATS[3],
            'fill': self.fillGreen
        }))
        return OrderedDict(meta)

    def process(self):
        count = 1
        current_row = self.Meta.content_start_row
        self.create_head()

        for k, v in self.data.total.items():
            self.fill_row(count, k, v, row=current_row)
            count += 1
            current_row += 1
        return

    @property
    def file_name(self):
        name = '{}_{}_{}_{}'.format(
            self.data.start.strftime('%Y-%b'),
            self.data.end.strftime('%Y-%b'),
            self.Meta.file_name, self.member_group)
        return '{}.xlsx'.format(name)
