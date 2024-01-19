from collections import OrderedDict

from openpyxl.styles.numbers import FORMAT_NUMBER_COMMA_SEPARATED1

from core.mixin import MonthMixIn
from core.report.excel import ExcelTemplateReport
from .summary_week_commission import SummaryWeekCommission


class ExcelSummaryWeekCommission(MonthMixIn, ExcelTemplateReport):
    """
    a class for generate excel object that represent summary of each member's week commission.
    Excel object can save to an excel file or response via http request file.
    This class inherit class `ExcelTemplateReport` also you can use a method in ExcelTemplateReport or overwrite method

    Attributes:
        start (:obj:`date`): a start date of interested data.
        end (:obj:`date`): a end date of interested data.
        get_type (str): time period option [daily, monthly, quarter, yearly]
        mcode (str): (Optional) member code of interested person
        honor (list): (Optional) member honor of interested group's member
    """
    template_file = './templates/report/commission/summary_week_commission.xlsx'

    class Meta:
        title = 'Week Commission'
        file_name = 'week_commission'
        head_file = 'Week commission'
        content_start_col = 1
        content_start_row = 8
        head_start_col = 4
        head_start_row = 6
        head = {'fields': []}

    def __init__(self, *args, **kwargs):
        super(ExcelSummaryWeekCommission, self).__init__(*args, **kwargs)
        self.include_resale = kwargs.get('resale', False)
        self.report = SummaryWeekCommission(*args, **kwargs)

    def create_head(self):
        """
        a method to process data to create Header of table in sheet of excel object
        """
        diff = self.month_diff_range(self.report.end, self.report.start)
        start_row = self.Meta.head_start_row
        start_col = self.Meta.head_start_col
        header = []
        for x in diff:
            header.append('{}'.format(x.strftime('%Y-%b')))

        self.work_sheet['E4'] = self.report.start.strftime('%Y-%m-%d')
        self.work_sheet['G4'] = self.report.end.strftime('%Y-%m-%d')
        self.work_sheet['E5'] = self.report.get_type
        super(ExcelSummaryWeekCommission, self).create_head()
        for x in header:
            self.work_sheet.cell(row=start_row, column=start_col).alignment = self.style['align_center']
            self.work_sheet.cell(row=start_row, column=start_col).value = x
            self.work_sheet.cell(row=start_row + 1, column=start_col).value = 'fast_bonus'
            self.work_sheet.cell(row=start_row + 1, column=start_col + 1).value = 'ws_bonus'

            if self.include_resale:
                self.work_sheet.cell(row=start_row + 1, column=start_col + 2).value = 'resale'
                self.work_sheet.cell(row=start_row + 1, column=start_col + 3).value = 'total'
                self.work_sheet.merge_cells(start_row=start_row,
                                            start_column=start_col,
                                            end_row=start_row,
                                            end_column=start_col + 3)
                start_col += 4

            else:
                self.work_sheet.cell(row=start_row + 1, column=start_col + 2).value = 'total'
                self.work_sheet.merge_cells(start_row=start_row,
                                            start_column=start_col,
                                            end_row=start_row,
                                            end_column=start_col + 2)
                start_col += 3
            self.work_sheet.cell(row=start_row + 1, column=start_col + 1).value = 'AVG'
            self.work_sheet.cell(row=start_row + 1, column=start_col + 1).value = 'Total'
        return

    def build_row_meta(self, row_index, data, **kwargs):
        """
        a method for index data with cell row

        :param row_index: (int) index of sheet row

        :param data: (:obj:`dictionary`) a content write in row

        :return: (:obj:`dictionary`)
        """
        row = kwargs.get('row')
        current_col = self.Meta.content_start_col + 3
        meta = [
            ('no', {'data': row_index, 'alignment': self.style['align_center']}),
            ('mcode', {'data': data['mcode'], 'alignment': self.style['align_center']}),
            ('name', {'data': data['name'], 'alignment': self.style['align_center']}),
        ]
        diff = self.month_diff_range(self.report.end, self.report.start)
        ws_cell = []
        total_cell = []
        count = 0
        for x in diff:
            month = x.strftime('%Y-%b')
            current_month = data['data'].get(month, None)
            if current_month is None:
                meta.append(('{}-FS-1'.format(month), {'data': 0,
                                                       'alignment': self.style['align_right'],
                                                       'number_format': FORMAT_NUMBER_COMMA_SEPARATED1
                                                       }))
                meta.append(('{}-WS-1'.format(month), {'data': 0,
                                                       'alignment': self.style['align_right'],
                                                       'number_format': FORMAT_NUMBER_COMMA_SEPARATED1
                                                       }))
                if self.include_resale:
                    meta.append(('{}-RS-1'.format(month), {'data': 0,
                                                           'alignment': self.style['align_right'],
                                                           'number_format': FORMAT_NUMBER_COMMA_SEPARATED1
                                                           }))
                    ws_cell.append(self.get_string_cell((current_col + (count * 4) + 3), row))
                    total_cell.append(self.get_string_cell((current_col + (count * 4) + 3), row))
                else:
                    ws_cell.append(self.get_string_cell((current_col + (count * 3) + 2), row))
                    total_cell.append(self.get_string_cell((current_col + (count * 3) + 2), row))
                meta.append(('{}-TS-1'.format(month), {'data': 0,
                                                       'alignment': self.style['align_right'],
                                                       'number_format': FORMAT_NUMBER_COMMA_SEPARATED1
                                                       }))
            else:
                meta.append(('{}-FS-1'.format(month), {'data': current_month.get('fast', 0),
                                                       'alignment': self.style['align_right'],
                                                       'number_format': FORMAT_NUMBER_COMMA_SEPARATED1
                                                       }))
                meta.append(('{}-WS-1'.format(month), {'data': current_month.get('ws', 0),
                                                       'alignment': self.style['align_right'],
                                                       'number_format': FORMAT_NUMBER_COMMA_SEPARATED1
                                                       }))
                if self.include_resale:
                    meta.append(('{}-RS-1'.format(month), {'data': current_month.get('resale', 0),
                                                           'alignment': self.style['align_right'],
                                                           'number_format': FORMAT_NUMBER_COMMA_SEPARATED1
                                                           }))
                    ws_cell.append(self.get_string_cell((current_col + (count * 4) + 3), row))
                    total_cell.append(self.get_string_cell((current_col + (count * 4) + 3), row))
                else:
                    ws_cell.append(self.get_string_cell((current_col + (count * 3) + 2), row))
                    total_cell.append(self.get_string_cell((current_col + (count * 3) + 2), row))

                meta.append(('{}-TS-1'.format(month), {'data': current_month.get('total', 0),
                                                       'alignment': self.style['align_right'],
                                                       'number_format': FORMAT_NUMBER_COMMA_SEPARATED1
                                                       }))

            count += 1

        meta.append(('{}-TS', {'data': '=AVERAGE({})'.format(','.join(ws_cell)),
                               'alignment': self.style['align_right'],
                               'number_format': FORMAT_NUMBER_COMMA_SEPARATED1
                               }))
        meta.append(('{}-TS1', {'data': '=SUM({})'.format(','.join(total_cell)),
                                'alignment': self.style['align_right'],
                                'number_format': FORMAT_NUMBER_COMMA_SEPARATED1
                                }))
        return OrderedDict(meta)

    def process_data(self):
        """
        a method to process data in excel object by call method `create_head` and `fill_row`
        """
        count = 1
        current_row = self.Meta.content_start_row
        self.create_head()
        for k, v in self.report.total.items():
            self.fill_row(count, v, row=current_row)
            count += 1
            current_row += 1

    @property
    def file_name(self):
        """
        a method to named a excel object

        :return: (str): excel object's file name
        """
        name = '{}'.format(
            self.Meta.file_name + '_' + self.report.start.strftime(
                '%Y-%m-%d') + '_' + self.report.end.strftime('%Y-%m-%d'))
        return '{}.xlsx'.format(name)
