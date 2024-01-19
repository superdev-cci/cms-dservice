from datetime import datetime, date
from collections import OrderedDict
from django.db.models import Q
from openpyxl.styles.numbers import FORMAT_NUMBER_COMMA_SEPARATED1
from core.report.excel import ExcelTemplateReport
from ecommerce.models import SaleInvoice
from branch.models import Branch

month_th = {'1': 'มกราคม', '2': 'กุมภาพันธ์', '3': 'มีนาคม', '4': 'เมษายน',
            '5': 'พฤษภาคม', '6': 'มิถุนายน', '7': 'กรกฎาคม', '8': 'สิงหาคม',
            '9': 'กันยายน', '10': 'ตุลาคม', '11': 'พฤศจิกายน', '12': 'ธันวาคม'}


class SaleInvoiceForAccountingReport(ExcelTemplateReport):
    """
    a class for create an excel file for Accounting department
    This class inherit class `ExcelTemplateReport` also you can use a method in ExcelTemplateReport or overwrite method

    Attributes:
        start (:obj:`date`): a start date of interested data.
        end (:obj:`date`): a end date of interested data.
        branch_target (str): branch code
    """
    template_file = './templates/report/ecommerce/sale_invoice_for_accounting.xlsx'

    class Meta:
        title = 'Sale Invoice for Accounting Report'
        file_name = 'sale_invoice_for_accounting'
        head_file = 'Sale Invoice for Accounting Report'
        head = {
            'fields': []
        }
        content_start_col = 1
        content_start_row = 11

    def __init__(self, **kwargs):
        super(SaleInvoiceForAccountingReport, self).__init__(**kwargs)
        self.start = kwargs.get('start', None)
        if isinstance(self.start, str):
            self.start = datetime.strptime(self.start, '%Y-%m-%d').date()
        elif self.start is None:
            self.start = date.today()

        self.end = kwargs.get('end', None)
        if isinstance(self.end, str):
            self.end = datetime.strptime(self.end, '%Y-%m-%d').date()
        elif self.end is None:
            self.end = date.today()

        self.branch_target = kwargs.get('branch', None)
        self.query_dict = self.get_query_dict()
        self.all_branch = {x.code: x for x in Branch.objects.all()}

    def create_head(self, sheet, branch_code, *args):
        """
        a method to process data to create Header of table in excel

        :param sheet: worksheet in excel file

        :param branch_code: a code that identify a branch
        """
        b = Branch.objects.get(inv_code=branch_code)
        sheet['B3'] = "เดือนภาษี " + month_th[str(self.end.month)]
        sheet[
            'D6'] = b.address + " ตำบล/แขวง " + b.sub_district + " อำเภอ/เขต " + b.district + " จังหวัด " + b.province + " " + b.post_code
        sheet['E7'] = "สาขาที่ " + b.branch_number

    def build_row_meta(self, row_index, itm, **kwargs):
        """
        a method for index data with cell row

        :param row_index: index of sheet row

        :param itm: (:obj:`dictionary`) a content write in row

        :return: (:obj:`dictionary`)
        """
        meta = OrderedDict([
            ('no', {'data': row_index, 'alignment': self.style['align_center']}),
            ('date', {'data': itm.sadate, 'alignment': self.style['align_center']}),
            ('bill_number', {'data': itm.sano}),
            ('client_code', {'data': itm.mcode}),
            ('client_name', {'data': itm.name_t}),
            ('client_tax', {'data': itm.member_id_card}),
            ('client_branch_num', {'data': ''}),
            ('client_branch_name', {'data': ''}),
            ('total_not_vat', {'data': '=K' + str(self.wb.active.max_row) + '-J' + str(self.wb.active.max_row),
                               'alignment': self.style['align_right'],
                               'number_format': FORMAT_NUMBER_COMMA_SEPARATED1}),
            ('vat', {'data': '=K' + str(self.wb.active.max_row) + '*7/107', 'alignment': self.style['align_right'],
                     'number_format': FORMAT_NUMBER_COMMA_SEPARATED1}),
            ('total', {'data': itm.total, 'alignment': self.style['align_right'],
                       'number_format': FORMAT_NUMBER_COMMA_SEPARATED1}),
            ('remark', {'data': itm.remark}),
        ])
        return meta

    def get_query_dict(self):
        """
        a method for process an information of SaleInvoice

        :return: (:obj:`dictionary`)
        """
        query_dict = {}
        if self.branch_target:
            query_dict[self.branch_target] = SaleInvoice.objects.filter(sadate__range=(self.start, self.end),
                                                                        inv_code=self.branch_target,
                                                                        bill_state='CM',
                                                                        cancel=0) \
                .filter(~Q(sa_type='Z'), ~Q(total=0)) \
                .select_related('member').order_by('sano')
        else:
            for br in Branch.objects.all().values_list('inv_code', flat=True):
                query_dict[br] = SaleInvoice.objects.filter(sadate__range=(self.start, self.end),
                                                            cancel=0,
                                                            bill_state='CM',
                                                            inv_code=br).filter(~Q(sa_type='Z'), ~Q(total=0)) \
                    .select_related('member').order_by('sano')
        return query_dict

    def fill_row(self, *args, **kwargs):
        """
        a method for write data to row with row meta index
        """
        row = kwargs.get('row')
        current_col = self.Meta.content_start_col
        cell = self.get_cell(current_col, row)
        cell.border = self.style['full_border']
        meta = self.build_row_meta(*args, **kwargs)
        for k, v in meta.items():
            cell = self.fill_data(cell, v)

    def process_data(self):
        """
        a method to process data in excel object by call method `create_head` and `fill_row`
        """
        source = self.wb.active
        self.create_head(source, 'BKK01')
        for nm in ['CRI01', 'HY01', 'KL01', 'STHQ']:
            target = self.wb.copy_worksheet(source)
            target.title = nm.split('0')[0]
            self.create_head(target, nm)
        for key in self.query_dict:
            self.wb.active = self.wb.worksheets.index(self.wb[key.split('0')[0]])
            count = 1
            current_row = self.wb.active.max_row
            for item in self.query_dict[key]:
                branch = self.all_branch.get(item.inv_code)
                if branch:
                    item.branch_num = branch.branch_number
                else:
                    item.branch_num = '-'
                if item.cancel == 1:
                    item.total = 0.0
                if item.member:
                    item.member_id_card = item.member.id_card
                else:
                    item.member_id_card = '-'
                self.fill_row(count, item, row=current_row)
                count += 1
                current_row += 1

    @property
    def file_name(self):
        """
        a method to named a excel object

        :return: (str): excel object's file name
        """
        name = '{}'.format(self.Meta.file_name)
        return '{}.xlsx'.format(name)
