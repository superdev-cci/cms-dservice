from commission.models import WeekCommission, WeekRound
from member.models import Member
from trip.models import Trip
from openpyxl import load_workbook
from statistics import median, stdev


class TripAnalyze(object):
    class Meta:
        trip_code = "2019CNX"
        path_excel = "/home/jew/Downloads/trip_2019CNX_2019-12-01.xlsx"

    def __init__(self, *args, **kwargs):
        super(TripAnalyze, self).__init__(*args, **kwargs)

    @property
    def fill_data_analyze(self):
        path_file = self.Meta.path_excel.split(".")
        filename = path_file[0]+"_with_weekcommission."+path_file[1]
        instance_trip = Trip.objects.get(code=self.Meta.trip_code)
        fdate_list = WeekRound.objects.filter(
            fdate__range=(instance_trip.start, instance_trip.end)).values_list("fdate", flat=True)
        wb = load_workbook(self.Meta.path_excel)
        ws = wb.active
        for i in range(ws.max_row):
            member = ws['A{}'.format(i + 1)].value
            ws['B{}'.format(i + 1)] = Member.objects.get(mcode=member).distributor_date.strftime("%Y-%m-%d")
            start_col = 3
            commission_list = []
            for fd in fdate_list:
                instance = WeekCommission.objects.filter(fdate=fd, mcode=member).values_list("ws_bonus", flat=True)
                if instance:
                    ws.cell(column=start_col, row=(i + 1), value=float(instance[0]))
                    commission_list.append(float(instance[0]))
                else:
                    ws.cell(column=start_col, row=(i + 1), value=float(0))
                    commission_list.append(0.0)
                start_col += 1
            ws.cell(column=(start_col + 1), row=(i + 1), value=float(median(commission_list)))
            ws.cell(column=(start_col + 2), row=(i + 1), value=float(stdev(commission_list)))
            ws.cell(column=(start_col + 3), row=(i + 1), value=float(sum(commission_list)))
        wb.save(filename)
