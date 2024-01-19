import datetime
import numbers
from django.db.models import Sum, Avg, Q, Count
from openpyxl import Workbook
from openpyxl import load_workbook

from core.mixin import MonthMixIn
from event.models import Attendee
from member.models import Member
from trip.models import Trip, TripApplication
from system_log.models import LogTravelCredit
from django.db.models.functions import TruncMonth, TruncDay, TruncYear, TruncQuarter
import statistics


def main():
    wb = load_workbook('/Users/saintent/Desktop/SG_TRIP.xlsx')
    ws = wb['SG']
    sg_pool = {}
    for i in range(ws.max_row):
        member = ws['B{}'.format(i + 1)].value
        right = ws['J{}'.format(i + 1)].value
        if member is None:
            continue
        elif right is None:
            continue
        # print('{} : -> {}'.format(member, right))
        sg_pool[member] = right

    select_trip = Trip.objects.get(name='Singapore')
    members = {x.mcode: x for x in Member.objects.filter(mcode__in=list(sg_pool.keys()))}
    create_pool = []
    for k, v in sg_pool.items():
        create_pool.append(TripApplication(member=members[k],
                                           trip=select_trip,
                                           register_date=select_trip.end,
                                           balance_use=select_trip.balance * v,
                                           seat=v
                                           ))
        print(k, v)
    TripApplication.objects.bulk_create(create_pool)
    return


def main_cnx():
    wb = load_workbook('/Users/saintent/Desktop/CNX_TRIP.xlsx')
    ws = wb['CNX']
    sg_pool = {}
    for i in range(ws.max_row):
        member = ws['B{}'.format(i + 1)].value
        right = ws['E{}'.format(i + 1)].value
        if member is None:
            continue
        elif right is None:
            continue
        # print('{} : -> {}'.format(member, right))
        sg_pool[member] = right

    select_trip = Trip.objects.get(code='2018CNX')
    members = {x.mcode: x for x in Member.objects.filter(mcode__in=list(sg_pool.keys()))}
    create_pool = []
    for k, v in sg_pool.items():
        create_pool.append(TripApplication(member=members[k],
                                           trip=select_trip,
                                           register_date=select_trip.end,
                                           balance_use=select_trip.balance * v,
                                           seat=v
                                           ))
        print(k, v)
    TripApplication.objects.bulk_create(create_pool)
    return


def main_bj():
    wb = load_workbook('/Users/saintent/Desktop/BJ_TRIP.xlsx')
    ws = wb['BJ']
    bj_pool = {}
    for i in range(ws.max_row):
        member = ws['A{}'.format(i + 1)].value
        right = ws['D{}'.format(i + 1)].value
        if member is None:
            continue
        elif right is None:
            continue
        # print('{} : -> {}'.format(member, right))
        bj_pool[member] = right

    select_trip = Trip.objects.get(code='2018BJ')
    members = {x.mcode: x for x in Member.objects.filter(mcode__in=list(bj_pool.keys()))}
    create_pool = []
    for k, v in bj_pool.items():
        create_pool.append(TripApplication(member=members[k],
                                           trip=select_trip,
                                           register_date=select_trip.end,
                                           balance_use=select_trip.balance * v,
                                           seat=v
                                           ))
        print(k, v)
    TripApplication.objects.bulk_create(create_pool)
    return


def main_at():
    wb = load_workbook('/Users/saintent/Desktop/AT_TRIP.xlsx')
    ws = wb['AT']
    bj_pool = {}
    for i in range(ws.max_row + 1):
        member = ws['B{}'.format(i + 2)].value
        right = ws['D{}'.format(i + 2)].value
        if member is None:
            continue
        elif right is None:
            continue
        # print('{} : -> {}'.format(member, right))
        bj_pool[member] = right

    select_trip = Trip.objects.get(code='2018AT')
    members = {x.mcode: x for x in Member.objects.filter(mcode__in=list(bj_pool.keys()))}
    create_pool = []
    for k, v in bj_pool.items():
        create_pool.append(TripApplication(member=members[k],
                                           trip=select_trip,
                                           register_date=select_trip.end,
                                           balance_use=select_trip.balance * v,
                                           seat=v
                                           ))
        print(k, v)
    TripApplication.objects.bulk_create(create_pool)
    return


def main_load():
    wb = load_workbook('/Users/saintent/Desktop/CNX2019-summary.xlsx')
    ws = wb.active
    scan_pool = {}
    total = 0
    for i in range(ws.max_row + 1):
        member = ws['B{}'.format(i + 2)].value
        right = ws['E{}'.format(i + 2)].value
        try:
            if member is None:
                continue
            elif right is None:
                continue
            elif right == 0 or right == ' ':
                continue
            # print('{} : -> {}'.format(member, right))
            total += int(right)
            scan_pool[member] = right
        except Exception as e:
            continue

    print('Total -> {}'.format(total))

    select_trip = Trip.objects.get(code='2019CNX')
    members = {x.mcode: x for x in Member.objects.filter(mcode__in=list(scan_pool.keys()))}
    create_pool = []
    for k, v in scan_pool.items():
        create_pool.append(TripApplication(member=members[k],
                                           trip=select_trip,
                                           register_date=select_trip.end,
                                           balance_use=select_trip.balance * v,
                                           seat=v
                                           ))
        print(k, v)
    TripApplication.objects.bulk_create(create_pool)
    return


def fix_balance_use():
    select_trip = Trip.objects.get(code='2019JP')
    query = TripApplication.objects.filter(trip=select_trip)
    for x in query:
        last = TripApplication.objects.filter(member=x.member,
                                              trip__trip_type='OS',
                                              ).filter(~Q(trip__code='2019JP')).count()
        if last == 0:
            new_balance = x.seat * (select_trip.balance - select_trip.balance_discount)
            print(x, 'OS : New -> {}'.format(new_balance))
            x.balance_use = new_balance
            x.save()
        else:
            print(x)


def fix_adjust_consume():
    wb = load_workbook('/Users/saintent/Desktop/trip_2019NR_2019-06-01.xlsx')
    ws = wb.active

    for i in range(ws.max_row + 1):
        member = ws['B{}'.format(i + 2)].value
        consume = ws['G{}'.format(i + 2)].value
        if member is not None and consume is not None:
            if isinstance(consume, numbers.Number):
                if consume == 0 or consume == 190000 or consume == 380000:
                    pass
                else:
                    print('{} : -> {}'.format(member, consume))
                    instance = TripApplication.objects.filter(member__mcode=member, trip__code='2019NR').first()
                    if instance is not None:
                        instance.previous_use = consume
                        instance.save()

    return


def test_hnn():
    wb = load_workbook('/Users/saintent/Desktop/trip_2020HHN_2020-02-22.xlsx')
    ws = wb.active
    count = 0
    for i in range(ws.max_row):
        member = ws['B{}'.format(i + 1)].value
        name = ws['C{}'.format(i + 1)].value
        balance = ws['E{}'.format(i + 1)].value
        count += 1
        event_attendant = Attendee.objects.filter(event__date__gte='2020-02-01',
                                                  members__mcode=member)
        m_instance = Member.objects.get(mcode=member)
        start = datetime.datetime.strptime('2019-01-01', '%Y-%m-%d').date()
        end = datetime.datetime.strptime('2020-01-01', '%Y-%m-%d').date()

        if start > m_instance.distributor_date:
            diff = MonthMixIn.month_diff_range(end, start)
        else:
            diff = MonthMixIn.month_diff_range(end, m_instance.distributor_date)
        pool = {}
        count = 0
        for x in diff:
            pool[x.strftime('%Y-%b')] = 0

        query = Member.objects.filter(sp_code=member, distributor_date__gt='2019-01-01', status_terminate=0) \
            .annotate(time=TruncMonth('distributor_date'), new_member=Count('level')) \
            .values('time', 'new_member').order_by('time')
        # sponsor = {x['level']: x['pos'] for x in query}
        for x in query:
            pool[x['time'].strftime('%Y-%b')] = x['new_member']

        array_data = [v for v in pool.values()]

        print(count, member, name, balance, event_attendant.count(), statistics.mean(array_data), statistics.median(array_data))

        ws['N{}'.format(i + 1)].value = event_attendant.count()
        ws['O{}'.format(i + 1)].value = statistics.mean(array_data)
        ws['P{}'.format(i + 1)].value = statistics.median(array_data)
        # ws['Q{}'.format(i + 1)].value = sponsor.get('DIS', 0)

    wb.save('/Users/saintent/Desktop/trip_2020HHN_2020-02-22.xlsx')


def load_tc_ask():
    base_balance = 20000
    # base_balance = 3750
    wb = load_workbook('/Users/saintent/Desktop/TC-Add.xlsx')
    ws = wb.active
    count = 0
    fail_pool = []
    logs = []
    for i in range(ws.max_row):
        try:
            member = ws['B{}'.format(i + 1)].value
            name = ws['C{}'.format(i + 1)].value
            right = int(ws['D{}'.format(i + 1)].value)
        except Exception as e:
            continue

        count += 1
        try:
            member_query = Member.objects.get(mcode=member)
            logs.append(LogTravelCredit(
                mcode=member, inv_code='System', sadate=datetime.date.today(), satime="11:00:00",
                sano='CREDITIN', value_in=(base_balance*right), value_out=0, total=(base_balance*right), uid='System',
                sa_type='',
                value_option="Monthly in"
            ))
            member_query.tc_value += base_balance * right
            member_query.save()
            print(member, name, right, (base_balance * right))
        except Exception as e:
            fail_pool.append(member)

    LogTravelCredit.objects.bulk_create(logs)



