from trip.models import Trip
from trip.models import TripApplication
from commission.models import WeekCommission
from member.models import Member
from dateutil.relativedelta import relativedelta
from openpyxl import Workbook, styles, load_workbook
from commission.report.summary.summary_month_commission import SummaryMonthCommission
from commission.report.summary.excel_summary_month_commission import ExcelSummaryMonthCommission


def get_cumulative(start, end, member=None):
    if member is None:
        cumulative = WeekCommission.get_cumulative_ws(start, end, member.code)
    else:
        cumulative = WeekCommission.get_cumulative_ws(start, end, member)
    balance = cumulative['total_balance']
    if balance is None:
        balance = 0
    return float(balance)


def get_maldives_consume():
    maldives = Trip.objects.get(code='2019NR')
    alaska = Trip.objects.get(code='AR2019')
    wb = Workbook()
    ws = wb.active
    start_overlap = alaska.start - relativedelta(days=1)
    row = 1
    for x in TripApplication.objects.filter(trip=maldives).select_related('member'):
        balance = get_cumulative(maldives.start, start_overlap, x.member)
        alaska_balance = get_cumulative(alaska.start, alaska.end, x.member)
        total_consume = int(balance - x.balance_use - x.previous_use)
        overlap_balance = get_cumulative(alaska.start, maldives.end, x.member)

        if total_consume >= 0:
            total_consume = 0
        print('Member {} : Total : {} -- Use : {} : Current : {} / {}  ---> {}'.format(
            x.member, balance, total_consume, x.balance_use, x.previous_use, overlap_balance))

        ws.cell(column=1, row=row, value=x.member.code)
        ws.cell(column=2, row=row, value=x.member.full_name)
        ws.cell(column=3, row=row, value=balance)
        ws.cell(column=4, row=row, value=overlap_balance)
        ws.cell(column=5, row=row, value=x.balance_use)
        ws.cell(column=6, row=row, value=x.previous_use)
        ws.cell(column=7, row=row, value=total_consume)
        ws.cell(column=8, row=row, value=alaska_balance)

        jp_query = TripApplication.objects.filter(trip__code='2019JP', member=x.member).first()
        if jp_query:
            ws.cell(column=9, row=row, value=jp_query.balance_use)
        else:
            ws.cell(column=9, row=row, value=0)
        row += 1

    wb.save('alaska.xlsx')


def get_matching():
    wb = load_workbook('/Users/saintent/Desktop/trip_AR2019_2020-06-01.xlsx')
    ws = wb.active
    member = []
    for i in range(ws.max_row):
        member_code = ws['B{}'.format(i + 1)].value
        member.append(member_code)

    report = SummaryMonthCommission(mcode=member, get_type='monthly', start='2019-05-01', end='2020-05-31')
    excel = ExcelSummaryMonthCommission(mcode=member, get_type='monthly', start='2019-05-01', end='2020-05-31')
    excel.process_data()
    excel.save_file()

    return report
