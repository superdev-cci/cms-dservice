from trip.models import Trip
from trip.models import TripApplication
from commission.models import WeekCommission
from member.models import Member
from dateutil.relativedelta import relativedelta
from openpyxl import Workbook, styles, load_workbook


def get_cumulative(start, end, member=None):
    if member is None:
        cumulative = WeekCommission.get_cumulative_ws(start, end, member.code)
    else:
        cumulative = WeekCommission.get_cumulative_ws(start, end, member)
    balance = cumulative['total_balance']
    if balance is None:
        balance = 0
    return float(balance)


def get_cnx_consume():
    jp = Trip.objects.get(code='2019JP')
    cnx = Trip.objects.get(code='2019CNX')
    wb = Workbook()
    ws = wb.active
    start_overlap = cnx.start - relativedelta(days=1)
    row = 1
    for x in TripApplication.objects.filter(trip=jp).select_related('member'):
        balance = get_cumulative(jp.start, start_overlap, x.member)
        cnx_balance = get_cumulative(cnx.start, cnx.end, x.member)
        total_consume = int(balance - x.balance_use - x.previous_use)
        overlap_balance = get_cumulative(cnx.start, cnx.end, x.member)

        if total_consume >= 0:
            total_consume = 0
        print('Member {} : Total : {} -- Use : {} : Current : {} / {}  ---> {}'.format(
            x.member, balance, total_consume, x.balance_use, x.previous_use, overlap_balance))

        ws.cell(column=1, row=row, value=x.member.code)
        ws.cell(column=2, row=row, value=x.member.full_name)
        ws.cell(column=3, row=row, value=balance)
        ws.cell(column=4, row=row, value=overlap_balance)
        ws.cell(column=5, row=row, value=x.balance_use)
        ws.cell(column=6, row=row, value=x.previous_use)
        ws.cell(column=7, row=row, value=total_consume)
        ws.cell(column=8, row=row, value=cnx_balance)
        row += 1
    wb.save('cnx.xlsx')
