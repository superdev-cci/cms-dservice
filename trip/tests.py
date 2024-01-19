from django.test import TestCase

from commission.models import MonthCommission
from trip.functions.trip_calculator import TripCalculator
import datetime
from django.db.models import Sum, Avg
from openpyxl import Workbook
from openpyxl import load_workbook
from member.models import Member
from trip.models import Trip, TripApplication
from member.models import Member
from system_log.models import LogTravelCredit
from trip.models import TravelPointUseStatement
from trip.functions import trip_pt_cal


# Create your tests here.
def test_cal():
    x = TripCalculator(member='TH8960571')
    x.find_active_trip()


def main():
    wb = load_workbook('/Users/saintent/Downloads/trip_2019NR_2019-04-08.xlsx')
    ws = wb.active
    member_pool = []
    for i in range(ws.max_row):
        member = ws['B{}'.format(i + 1)].value
        if member and member != 'Member code':
            member_pool.append(member)
            oversea = TripApplication.objects.filter(member__mcode=member, trip__trip_type='OS').count()
            if oversea:
                ws['I{}'.format(i + 1)].value = '-'
            else:
                ws['I{}'.format(i + 1)].value = 'NEW'
            # print(member)
            print(member)

    wb.save('/Users/saintent/Downloads/trip_2019NR_2019-04-08.xlsx')
    return


def revert_tc():
    queryset = LogTravelCredit.objects.filter(sadate='2020-10-02')
    for x in queryset:
        print(x.mcode, x.value_out, x.value_option)
        Member.objects.filter(mcode=x.mcode).update(tc_value=x.value_out)


def check_report_pt():
    t = Trip.objects.get(code="PT2021")
    pd = trip_pt_cal.check_report_pt(t)
    return pd


def check_report_nl():
    trip = Trip.objects.get(code="2020NR")
    rp = TravelPointUseStatement.objects.filter(trip=trip) \
        .values('member__mcode', 'member__name_t') \
        .annotate(total_gold=Sum('gold_coin'),
                  total_silver=Sum('silver_coin')) \
        .filter(total_silver__gte=100) \
        .order_by('member')

    for x in rp:
        m = Member.objects.get(mcode=x.get('member__mcode'))
        matching_count = MonthCommission.objects.filter(fdate__range=(trip.start, trip.end),
                                                        mcode=m.code, dmbonus__gte=0)

        print("{},{},{},{},{}".format(x['member__mcode'],
                                      x['member__name_t'],
                                      matching_count.count(),
                                      x['total_gold'],
                                      x['total_silver'],
                                      )
              )
