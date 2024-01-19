from datetime import datetime
from openpyxl import Workbook, styles
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
from openpyxl.utils import get_column_letter
from openpyxl.writer.excel import save_virtual_workbook
from openpyxl.styles import colors

thin = Side(border_style="thin", color=colors.BLACK)
double = Side(border_style="double", color=colors.BLACK)
border = Border(top=thin, left=thin, right=thin, bottom=thin)
alc = Alignment(horizontal="center", vertical="center")


class ExcelReportBase(object):
    style = {
        'full_border': border,
        'align_center': alc,
        'align_right': Alignment(horizontal="right", vertical="center"),
        'align_left': Alignment(horizontal="left", vertical="center")
    }

    def __init__(self, *args, **kwargs):
        self.wb = Workbook()
        self.work_sheet = self.wb.active
        self.criteria = None

    def create_title(self):
        Meta = self.Meta
        self.work_sheet.title = Meta.sheet_name
        # Title
        self.work_sheet.merge_cells(start_row=1, start_column=1, end_row=2, end_column=5)
        self.work_sheet.cell(column=1, row=1, value=Meta.title)
        self.work_sheet.cell(column=1, row=1).style = 'Title'
        # Date create
        self.work_sheet.merge_cells(start_row=3, start_column=1, end_row=3, end_column=5)
        self.work_sheet.cell(column=1, row=3).value = 'Date create : {}'.format(
            datetime.now().strftime('%Y-%M-%d %H:%M:%S'))

    def create_head(self, *args):
        Meta = self.Meta
        start_row = Meta.head_start_row
        start_col = Meta.head_start_col
        idx = 0
        if 'merge-h' in Meta.head:
            merge_h = Meta.head.get('merge-h', 1)
            for x in Meta.head['fields']:
                self.work_sheet.merge_cells(start_row=start_row,
                                            start_column=start_col + idx,
                                            end_row=start_row + merge_h,
                                            end_column=start_col + idx)
                ch = ord('A')
                c_range = '{}{}:{}{}'.format(chr(ch + idx), start_row, chr(ch + idx), start_row + merge_h)
                self.style_range(self.work_sheet,
                                 c_range,
                                 self.style['full_border'],
                                 self.style['align_center'])
                self.work_sheet.column_dimensions['{}'.format(chr(ch + idx))].width = 15
                # self.ws.cell(row=start_row, column=start_col).border = self.style['full_border']
                # self.ws.cell(row=start_row, column=start_col).alignment = self.style['align_center']
                idx += 1

        for x in Meta.head['fields']:
            self.work_sheet.cell(row=start_row, column=start_col).value = x
            start_col += 1

    def style_range(self, ws, cell_range, border=Border(), alignment=None, fill=None, font=None, ):

        top = Border(top=border.top)
        left = Border(left=border.left)
        right = Border(right=border.right)
        bottom = Border(bottom=border.bottom)

        first_cell = ws[cell_range.split(":")[0]]
        if alignment:
            ws.merge_cells(cell_range)
            first_cell.alignment = alignment

        rows = ws[cell_range]
        if font:
            first_cell.font = font

        for cell in rows[0]:
            cell.border = cell.border + top
        for cell in rows[-1]:
            cell.border = cell.border + bottom

        for row in rows:
            l = row[0]
            r = row[-1]
            l.border = l.border + left
            r.border = r.border + right
            if fill:
                for c in row:
                    c.fill = fill
        return

    def fill_color(self, cell, color):
        color = styles.colors.Color(rgb=color)
        fill_color = styles.fills.PatternFill(patternType='solid', fgColor=color)
        cell.fill = fill_color

    def get_cell(self, col, row):
        return self.wb.active.cell(column=col, row=row)

    def next_cell(self, cell, col, row):
        return self.wb.active[self.next_string_cell(cell, col, row)]

    def next_string_cell(self, cell, col, row):
        next_idx = cell.col_idx + col
        return '{}{}'.format(get_column_letter(next_idx), cell.row + row)

    def get_string_cell(self, col, row):
        return '{}{}'.format(get_column_letter(col), row)

    def get_string_cell_range(self, start_col, start_row, end_col, end_row):
        start_cell = self.get_string_cell(start_col, start_row)
        end_cell = self.get_string_cell(end_col, end_row)

        return '{}:{}'.format(start_cell, end_cell)

    def merge_cell(self, cell, column, row, border=Border(), alignment=None, fill=None, font=None):
        start = cell.coordinate
        end = self.next_string_cell(cell, column, row)
        cell_range = '{}:{}'.format(start, end)
        self.work_sheet.merge_cells(cell_range)
        self.style_range(self.work_sheet,
                         cell_range,
                         self.style['full_border'],
                         self.style['align_center'])

    def apply_header_style(self, cell):
        cell.border = self.style['full_border']
        cell.alignment = self.style['align_center']

    def apply_border(self, cell):
        cell.border = self.style['full_border']

    @property
    def response_file(self):
        return save_virtual_workbook(self.wb)

    def save_file(self, ):
        self.wb.save(self.file_name)

    @property
    def file_name(self):
        raise NotImplementedError('`file_name` must be implemented.')
