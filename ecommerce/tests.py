from datetime import date
from dateutil.relativedelta import relativedelta
from django.db.models import Sum
from django.test import TestCase
from ecommerce.report import SoldMonthlySummary, SoldDailySummary, \
    SoldItemDailySummary, SoldItemMonthlySummary, SoldMonthlyPaymentSummary
from ecommerce.models import Product, ProductClass, Promotion, SaleItem, SaleInvoice
from ecommerce.report.distributor import AgencyBoughtReport, FriendchiseBoughtReport, SalesMemberTypeReport
from ecommerce.report.sale import SaleItemSummaryReport
from openpyxl import Workbook, styles, load_workbook
from django.db import models
from member.models import Member


def get_promotion_sales():
    queryset = SaleItem.objects.filter(sadate__year=2017).values('pcode', 'pdesc') \
        .annotate(total_prices=Sum('amt'), total_qty=Sum('qty')) \
        .order_by('pcode', )
    count = 1
    start = 2
    wb = Workbook()
    ws = wb.active
    ws.cell(column=1, row=1, value='No.')
    ws.cell(column=2, row=1, value='Promotion code')
    ws.cell(column=3, row=1, value='Description')
    ws.cell(column=4, row=1, value='QTY')
    ws.cell(column=5, row=1, value='Prices')
    ws.cell(column=6, row=1, value='Start')
    ws.cell(column=7, row=1, value='End')
    ws.cell(column=8, row=1, value='Margin')
    for x in queryset:
        promotion = Promotion.objects.filter(pcode=x['pcode']).first()
        if promotion is None:
            continue

        start_date = SaleItem.objects.filter(sadate__year=2017, pcode=x['pcode']).order_by('sadate').first()
        end_date = SaleItem.objects.filter(sadate__year=2017, pcode=x['pcode']).order_by('sadate').last()
        print(x, promotion, start_date, end_date)
        if start_date is None or end_date is None:
            continue
        print('PCODE : {}, PDEST : {} --> Total amt: {} --> Total prices: {}, start={}, end={}'.format(x['pcode'],
                                                                                                       x['pdesc'], x[
                                                                                                           'total_prices'],
                                                                                                       x['total_qty'],
                                                                                                       start_date.sadate,
                                                                                                       end_date.sadate))
        ws.cell(column=1, row=start, value=count)
        ws.cell(column=2, row=start, value=x['pcode'])
        ws.cell(column=3, row=start, value=x['pdesc'])
        ws.cell(column=4, row=start, value=x['total_qty'])
        ws.cell(column=5, row=start, value=x['total_prices'])
        ws.cell(column=6, row=start, value=start_date.sadate)
        ws.cell(column=7, row=start, value=end_date.sadate)
        ws.cell(column=8, row=start, value=promotion.margin_net)
        start += 1
        count += 1

    wb.save('promotion_sale_2019.xlsx')

    return


# Get all of item to sales with margin
def main_test():
    start = 2
    product = {x.pcode: x for x in Product.objects.all()}
    promotion = {x.pcode: x for x in Promotion.objects.all()}
    x = SaleItemSummaryReport(start='2019-01-01', end='2019-12-31', get_type='yearly', only_value=True)
    wb = Workbook()
    ws = wb.active
    ws.cell(column=1, row=1, value='Product code')
    ws.cell(column=2, row=1, value='Description')
    ws.cell(column=3, row=1, value='QTY')
    ws.cell(column=4, row=1, value='Prices')
    ws.cell(column=5, row=1, value='Margin')

    pool = {}

    for item in x.total.values():
        for k, v in item.items():
            code = k
            print(k, v)
            # if k.startswith('LCCI'):
            #     code = k[1:]
            if code in pool:
                pool[code]['total_prices'] += v['total_prices']
                pool[code]['total_qty'] += v['total_qty']
            else:
                pool[code] = v

    for k, v in pool.items():
        ws.cell(column=1, row=start, value=k)
        if k in product:
            ws.cell(column=2, row=start, value=product[k].pdesc)
            ws.cell(column=5, row=start, value=product[k].margin_net)
        else:
            ws.cell(column=2, row=start, value=promotion[k].pdesc)
            ws.cell(column=5, row=start, value=promotion[k].margin_net)
        ws.cell(column=3, row=start, value=v['total_qty'])
        ws.cell(column=4, row=start, value=v['total_prices'])
        start += 1
    wb.save('sales_items.xlsx')
    return


def main():
    # x = SoldSummary(start='2018-10-01', end='2018-10-24')
    x = SoldMonthlySummary(start='2019-01-01', end='2019-12-31')
    # x = SoldMonthlySummary(start='2018-01-01', end='2018-10-24')
    # x = SoldItemDailySummary(start='2018-10-02', end='2018-10-04')
    # x = SoldItemMonthlySummary(start='2016-01-01', end='2016-12-31')
    # x = SoldMonthlyPaymentSummary(start='2018-01-02', end='2018-10-04')

    x.process_data()
    x.save_file()


def init_product():
    Product.objects.update(activated=False)
    Product.objects.filter(pcode__startswith='CCI').update(activated=True)
    Product.objects.filter(price=195).update(personel_price=300)
    Product.objects.filter(price=395).update(personel_price=540)
    Product.objects.filter(price=550).update(personel_price=950)
    Product.objects.filter(price=590).update(personel_price=980)
    Product.objects.filter(price=950).update(personel_price=1600)


def init_product_class():
    Product.objects.filter(price=195).update(product_class=ProductClass.objects.get(name='Tier3'))
    Product.objects.filter(price=395).update(product_class=ProductClass.objects.get(name='Tier4'))
    Product.objects.filter(price=550).update(product_class=ProductClass.objects.get(name='Tier1'))
    Product.objects.filter(price=590).update(product_class=ProductClass.objects.get(name='Tier2'))
    Product.objects.filter(price=950).update(product_class=ProductClass.objects.get(name='Tier5'))


def test_sale_group():
    m = Member.objects.get(mcode='TH0123456')
    query = SaleItem.objects.filter(sano_link__member__line_lft__gte=m.line_lft,
                                    sano_link__member__line_rgt__lte=m.line_rgt, sano_link__sa_type__in=('A', 'H'),
                                    sano_link__cancel=0)
    query = query.filter(~models.Q(sano_link__bill_state__in=('CA', 'OR')),
                         sano_link__sadate__range=('2019-03-01', '2019-03-31'))
    query = query.values('pcode', 'pdesc',
                         'sano_link__member__mcode',
                         'sano_link__member__name_t') \
        .annotate(total_buy=models.Sum('amt'), total_qty=models.Sum('qty')).order_by('pcode',
                                                                                     'sano_link__member__name_t')
    for x in query:
        print(x)


def test_ag_report():
    select_date = date.today() - relativedelta(months=1)
    ag = AgencyBoughtReport(date=select_date)
    ag.process_data()
    ag.save_file()


def test_fr_report():
    select_date = date.today() - relativedelta(months=1)
    ag = FriendchiseBoughtReport(date=select_date)
    ag.process_data()
    ag.save_file()


def get_sales_report():
    # fr = SalesMemberTypeReport(type=('FR', 'MF'))
    # fr.process_data()
    # fr.save_file()

    ag = SalesMemberTypeReport(type=('AG',))
    ag.process_data()
    ag.save_file()


def test_tomazinc():
    from ecommerce.report.excel.sale_item import MemberBoughtItemSummaryExcel, MemberBoughtItemCodeSummaryExcel
    x = MemberBoughtItemSummaryExcel(start='2019-01-01', end='2019-05-31', items=('CCI008',),
                                     get_type='monthly')
    # x = MemberBoughtItemCodeSummaryExcel(start='2019-04-28', end='2019-05-04', items=('CCI008',))
    x.process()
    x.save_file()


def test_relax():
    from ecommerce.report.excel.sale_item import MemberBoughtItemSummaryExcel, MemberBoughtItemCodeSummaryExcel
    x = MemberBoughtItemSummaryExcel(start='2018-01-01', end='2018-12-31', items=('CCI002',),
                                     get_type='yearly', merge=True)
    # x = MemberBoughtItemCodeSummaryExcel(start='2019-04-28', end='2019-05-04', items=('CCI008',))
    x.process()
    x.save_file()


def create_sales_with_sponsor_report():
    from ecommerce.report.excel.sale.sales_honor_group_with_sponsor_structure_excel import \
        SalesHonorGroupWithStructureExcelReport
    start_date = '2019-01-01'
    end_date = '2019-11-30'

    # x = SalesHonorGroupWithStructureExcelReport(start=start_date, end=end_date, group='CE')
    # x.process()
    # x.save_file()
    #
    # x = SalesHonorGroupWithStructureExcelReport(start=start_date, end=end_date, group='DE')
    # x.process()
    # x.save_file()

    x = SalesHonorGroupWithStructureExcelReport(start=start_date, end=end_date, group='EE')
    x.process()
    x.save_file()

    x = SalesHonorGroupWithStructureExcelReport(start=start_date, end=end_date, group='SE')
    x.process()
    x.save_file()
    #
    # x = SalesHonorGroupWithStructureExcelReport(start=start_date, end=end_date, group='PE')
    # x.process()
    # x.save_file()
    #
    # x = SalesHonorGroupWithStructureExcelReport(start=start_date, end=end_date, group='GL')
    # x.process()
    # x.save_file()
    #
    # x = SalesHonorGroupWithStructureExcelReport(start=start_date, end=end_date, group='SI')
    # x.process()
    # x.save_file()

    return

def fix_pro():
    wb = load_workbook('/Users/saintent/Desktop/edit_pro.xlsx')
    ws = wb.active
    sano_pool = []
    for i in range(ws.max_row):
        sano = ws['L{}'.format(i + 1)].value
        sano_pool.append(sano)
        # print(sano)

    query_set = SaleInvoice.objects.prefetch_related('items').filter(sano__in=sano_pool,
                                                                     sadate__gte='2020-10-01').select_related('member')
    for x in query_set:
        print(x.sano, x.member.mcode, x.sadate)
        total_pv = 0
        total_prices = 0
        for y in x.items.all():
            print("\t\t", y.pcode, y.pdesc, y.qty, y.amt, y.pv)
            if y.pcode == 'PCCIDS005':
                total_pv += y.qty * 250
                total_prices += y.qty * 2750
                # y.pcode = 'PCCIDS005'
                # y.pdesc = 'Olisa 5 กล่อง แถมฟรี 1 กล่อง'
                y.pv = y.qty * 250
                y.price = 2750
                y.customer_price = 2750
                y.amt = y.qty * 2750
                y.save()
            else:
                total_pv += y.pv * y.qty
                total_prices += y.amt
        x.total = total_prices
        x.pv = total_pv
        x.total_vat = float(total_prices) - float('{:.2f}'.format(total_prices * 100 / 107))
        x.total_net = float(total_prices) - float(x.total_vat)
        x.total_invat = x.total_net
        x.save()
        print("\t\tTotal : pv ", total_pv)
        print("\t\tTotal : prices ", total_prices, 'VAT', '{:.2f}'.format(total_prices * 100 / 107))
region_dict = {
    "1": "ภาคเหนือ",
    "2": "ภาคกลาง",
    "3": "ภาคตะวันออกเฉียงเหนือ",
    "4": "ภาคตะวันตก",
    "5": "ภาคตะวันออก",
    "6": "ภาคใต้",
}

province_dict = {
    "1": {"name": "กรุงเทพมหานคร", "region": 2},
    "2": {"name": "สมุทรปราการ", "region": 2},
    "3": {"name": "นนทบุรี", "region": 2},
    "4": {"name": "ปทุมธานี", "region": 2},
    "5": {"name": "พระนครศรีอยุธยา", "region": 2},
    "6": {"name": "อ่างทอง", "region": 2},
    "7": {"name": "ลพบุรี", "region": 2},
    "8": {"name": "สิงห์บุรี", "region": 2},
    "9": {"name": "ชัยนาท", "region": 2},
    "10": {"name": "สระบุรี", "region": 2},
    "11": {"name": "นครนายก", "region": 2},
    "12": {"name": "ชลบุรี", "region": 5},
    "13": {"name": "ระยอง", "region": 5},
    "14": {"name": "จันทบุรี", "region": 5},
    "15": {"name": "ตราด", "region": 5},
    "16": {"name": "ฉะเชิงเทรา", "region": 5},
    "17": {"name": "ปราจีนบุรี", "region": 5},
    "18": {"name": "สระแก้ว", "region": 5},
    "19": {"name": "นครราชสีมา", "region": 3},
    "20": {"name": "บุรีรัมย์", "region": 3},
    "21": {"name": "สุรินทร์", "region": 3},
    "22": {"name": "ศรีสะเกษ", "region": 3},
    "23": {"name": "อุบลราชธานี", "region": 3},
    "24": {"name": "ยโสธร", "region": 3},
    "25": {"name": "ชัยภูมิ", "region": 3},
    "26": {"name": "อำนาจเจริญ", "region": 3},
    "27": {"name": "หนองบัวลำภู", "region": 3},
    "28": {"name": "ขอนแก่น", "region": 3},
    "29": {"name": "อุดรธานี", "region": 3},
    "30": {"name": "เลย", "region": 3},
    "31": {"name": "หนองคาย", "region": 3},
    "32": {"name": "มหาสารคาม", "region": 3},
    "33": {"name": "ร้อยเอ็ด", "region": 3},
    "34": {"name": "กาฬสินธุ์", "region": 3},
    "35": {"name": "สกลนคร", "region": 3},
    "36": {"name": "นครพนม", "region": 3},
    "37": {"name": "มุกดาหาร", "region": 3},
    "38": {"name": "เชียงใหม่", "region": 1},
    "39": {"name": "ลำพูน", "region": 1},
    "40": {"name": "ลำปาง", "region": 1},
    "41": {"name": "อุตรดิตถ์", "region": 1},
    "42": {"name": "แพร่", "region": 1},
    "43": {"name": "น่าน", "region": 1},
    "44": {"name": "พะเยา", "region": 1},
    "45": {"name": "เชียงราย", "region": 1},
    "46": {"name": "แม่ฮ่องสอน", "region": 1},
    "47": {"name": "นครสวรรค์", "region": 2},
    "48": {"name": "อุทัยธานี", "region": 2},
    "49": {"name": "กำแพงเพชร", "region": 2},
    "50": {"name": "สุโขทัย", "region": 2},
    "51": {"name": "พิษณุโลก", "region": 2},
    "52": {"name": "พิจิตร", "region": 2},
    "53": {"name": "เพชรบูรณ์", "region": 2},
    "54": {"name": "สุพรรณบุรี", "region": 2},
    "55": {"name": "นครปฐม", "region": 2},
    "56": {"name": "สมุทรสาคร", "region": 2},
    "57": {"name": "สมุทรสงคราม", "region": 2},
    "58": {"name": "ตาก", "region": 4},
    "59": {"name": "ราชบุรี", "region": 4},
    "60": {"name": "กาญจนบุรี", "region": 4},
    "61": {"name": "เพชรบุรี", "region": 4},
    "62": {"name": "ประจวบคีรีขันธ์", "region": 4},
    "63": {"name": "นครศรีธรรมราช", "region": 6},
    "64": {"name": "กระบี่", "region": 6},
    "65": {"name": "พังงา", "region": 6},
    "66": {"name": "ภูเก็ต", "region": 6},
    "67": {"name": "สุราษฎร์ธานี", "region": 6},
    "68": {"name": "ระนอง", "region": 6},
    "69": {"name": "ชุมพร", "region": 6},
    "70": {"name": "สงขลา", "region": 6},
    "71": {"name": "สตูล", "region": 6},
    "72": {"name": "ตรัง", "region": 6},
    "73": {"name": "พัทลุง", "region": 6},
    "74": {"name": "ปัตตานี", "region": 6},
    "75": {"name": "ยะลา", "region": 6},
    "76": {"name": "นราธิวาส", "region": 6},
    "77": {"name": "บึงกาฬ", "region": 3}
}

# from ecommerce.models import SaleInvoice, SaleItem
# from django.db.models import Max, Min, Avg, Sum, Count
# target = "สุรินทร์"
#
# result = {}
# primary = SaleInvoice.objects.prefetch_related('items').filter(
#     sadate__range=("2019-12-01", "2019-12-31"),
#     cancel=0,
#     total__gt=10,
#     sa_type__in=["A", "H"]
# )
# qs1 = primary.filter(send=1).values("cprovinceid").annotate(
#     bill_count=Count("id"),
#     sum_total=Sum("total")
# ).order_by("-sum_total")
# qs2 = primary.filter(send=2).values("member__cprovinceid").annotate(
#     bill_count=Count("id"),
#     sum_total=Sum("total")
# ).order_by("-sum_total")
#
# for i in qs1:
#     pop = primary.filter(cprovinceid=i["cprovinceid"]).values_list("mcode", flat=True).order_by("mcode").distinct()
#     if i["cprovinceid"] not in result:
#         result[i["cprovinceid"]] = {
#             "sum_total": int(i["sum_total"]),
#             "bill_count": i["bill_count"],
#             "pop": len(pop)
#         }
#     else:
#         result[i["cprovinceid"]]["sum_total"] += int(i["sum_total"])
#         result[i["cprovinceid"]]["bill_count"] += i["bill_count"]
#         result[i["cprovinceid"]]["pop"] += len(pop)
#
# for i in qs2:
#     pop = primary.filter(send=2, member__provinceid=i["member__cprovinceid"]).values_list("mcode", flat=True).order_by("mcode").distinct()
#     if i["member__cprovinceid"] not in result:
#         result[i["member__cprovinceid"]] = {
#             "sum_total": int(i["sum_total"]),
#             "bill_count": i["bill_count"],
#             "pop": len(pop)
#         }
#     else:
#         result[i["member__cprovinceid"]]["sum_total"] += int(i["sum_total"])
#         result[i["member__cprovinceid"]]["bill_count"] += i["bill_count"]
#         result[i["member__cprovinceid"]]["pop"] += len(pop)
#
#
# from commission.models import PvTransfer
#
# target = "TH2469789"
# primary = PvTransfer.objects.filter(
#     uid=target,
#     sadate__range=("2019-12-01", "2019-12-31"),
#     cancel=0
# ).values("member__cprovinceid", "mcode").annotate(
#     bill_count=Count("id"),
#     sum_total=Sum("tot_pv")
# ).order_by("-sum_total")
