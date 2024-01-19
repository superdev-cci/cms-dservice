from openpyxl import load_workbook
from ecommerce.models import Product


def main():
    wb = load_workbook('/Users/saintent/Desktop/CCI/ProductMargin.xlsx')
    ws = wb.active
    for i in range(ws.max_row - 1):
        code = ws['A{}'.format(i + 2)].value
        cost = ws['G{}'.format(i + 2)].value
        if cost != 0 and code not in ["CCI050", ]:
            print(code, cost)
            Product.objects.filter(pcode=code).update(cost=cost)
