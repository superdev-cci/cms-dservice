from abc import ABC
from collections import OrderedDict
from datetime import date

from openpyxl.styles import PatternFill
from openpyxl.styles.numbers import FORMAT_NUMBER_COMMA_SEPARATED1, BUILTIN_FORMATS
from core.mixin import MonthMixIn
from core.report.excel import ExcelTemplateReport
from ecommerce.report.sale.sales_member_with_sponsor_structure import SalesMemberWithSponsorStructureJSONReport


class SalesHonorGroupWithStructureExcelReport(MonthMixIn, ExcelTemplateReport, ABC):
    template_file = './templates/report/ecommerce/member_sales.xlsx'
    fillRed = PatternFill("solid", fgColor='00f5b0af')
    fillYellow = PatternFill("solid", fgColor='00f4e694')
    fillGreen = PatternFill("solid", fgColor='00b4f198')

    class Meta:
        title = 'SALES WITH SPONSOR STRUCTURE'
        sheet_name = 'Summary'
        file_name = 'sales_with_sponsor_structure_summary'
        head_start_col = 5
        head_start_row = 7
        head = {
            'fields': []
        }
        content_start_col = 1
        content_start_row = 8

    def __init__(self, *args, **kwargs):
        super(SalesHonorGroupWithStructureExcelReport, self).__init__(*args, **kwargs)
        self.get_type = kwargs.get('get_type', None)
        self.member_group = kwargs.get('group', None)
        self.data = SalesMemberWithSponsorStructureJSONReport(**kwargs)

    def create_head(self):
        diff = self.month_diff_range(self.data.end, self.data.start)
        start_row = self.Meta.head_start_row
        start_col = self.Meta.head_start_col
        header = []
        for x in diff:
            header.append('{}'.format(x.strftime('%Y-%b')))
        header.append('TOTAL-SALES')

        self.work_sheet['C3'] = self.Meta.title
        self.work_sheet['D4'] = date.today().strftime('%d/%b/%Y')

        for x in header:
            self.work_sheet.cell(row=start_row, column=start_col).value = x
            start_col += 1
        return

    def build_row_meta(self, row_index, mcode, data, **kwargs):
        meta = [
            ('no', {'data': row_index, 'alignment': self.style['align_center']}),
            ('mcode', {'data': mcode, 'alignment': self.style['align_left']}),
            ('name', {'data': data['name'], 'alignment': self.style['align_left']}),
            ('honor', {'data': data['honor'], 'alignment': self.style['align_center']})
        ]
        diff = self.month_diff_range(self.data.end, self.data.start)
        total_sales = 0
        for x in diff:
            month = x.strftime('%Y-%b')
            current_month = data['data'].get(month, None)
            if current_month is None:
                meta.append(('{}-1'.format(month), {'data': 0, 'alignment': self.style['align_right']}))
            else:
                sales = current_month.get('sales', 0)
                total_sales += sales
                meta.append(('{}-1'.format(month), {
                    'data': sales,
                    'alignment': self.style['align_right'],
                    'number_format': BUILTIN_FORMATS[3]
                }))
        meta.append(('total-1', {
            'data': total_sales,
            'alignment': self.style['align_right'],
            'number_format': BUILTIN_FORMATS[3],
            'fill': self.fillGreen
        }))
        return OrderedDict(meta)

    def process(self):
        count = 1
        current_row = self.Meta.content_start_row
        self.create_head()

        for k, v in self.data.total.items():
            self.fill_row(count, k, v, row=current_row)
            count += 1
            current_row += 1
        return

    @property
    def file_name(self):
        name = '{}_{}_{}_{}'.format(
            self.data.start.strftime('%Y-%b'),
            self.data.end.strftime('%Y-%b'),
            self.Meta.file_name, self.member_group)
        return '{}.xlsx'.format(name)
