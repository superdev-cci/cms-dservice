from collections import OrderedDict
from datetime import date

from openpyxl.styles import PatternFill
from openpyxl.styles.numbers import FORMAT_NUMBER_COMMA_SEPARATED1, BUILTIN_FORMATS

from commission.report import FrPvTransferInReport
from core.mixin import MonthMixIn
from core.report.excel import ExcelTemplateReport
from ecommerce.report.sale import SalesAgFrSummaryReport
from member.models import Member


class SalesMemberTypeReport(ExcelTemplateReport, MonthMixIn):
    template_file = './templates/report/ecommerce/member_bought ag_fr.xlsx'
    margin = 5000
    fillRed = PatternFill("solid", fgColor='00f29a98')
    fillGreen = PatternFill("solid", fgColor='00b4f198')

    class Meta:
        title = 'Fr Qualified check'
        sheet_name = 'Summary'
        file_name = 'fg_check'
        head_file = 'Ag Qualified check'
        head_start_col = 4
        head_start_row = 6
        content_start_col = 1
        content_start_row = 7

    def __init__(self, *args, **kwargs):
        super(SalesMemberTypeReport, self).__init__(**kwargs)
        self.select_date = kwargs.get('date', date.today())
        self.start, self.end = self.get_year_range(self.select_date)
        self.bough_history = SalesAgFrSummaryReport(start=self.start, end=self.end, get_type='monthly',
                                                    type=kwargs.get('type', ('FR', 'MF')))
        self.select_type = kwargs.get('type', ('FR', 'MF'))

    def create_head(self):
        self.work_sheet['D4'] = 'Agency'
        self.work_sheet['D5'] = date.today().strftime('%d/%b/%Y')
        current_col = self.Meta.head_start_col
        cell = self.get_cell(current_col, self.Meta.head_start_row)
        cell.border = self.style['full_border']
        for x in self._calculate_month_range(self.start, self.end):
            current = x.strftime('%Y-%b')
            cell.value = current
            cell = self.next_cell(cell, 1, 0)
        cell.value = 'Total'

    def build_row_meta(self, row_index, data, **kwargs):
        row_meta = [
            ('no', {'data': row_index, 'alignment': self.style['align_center']}),
            ('mcode', {'data': data['code']}),
            ('name', {'data': data['name']}),
        ]
        month_range = self._calculate_month_range(self.start, self.end)
        pass_count = 0
        for x in month_range:
            current_month = x.strftime('%Y-%b')
            pv_month = data['data'].get(current_month, 0)
            total_pv = pv_month
            row_meta.append((current_month, {
                'data': total_pv,
                'number_format': BUILTIN_FORMATS[3],
                'alignment': self.style['align_right'],
            }))
        return OrderedDict(row_meta)

    def process_data(self):
        count = 1
        current_row = self.Meta.content_start_row
        self.create_head()
        total = self.bough_history.total
        merge_pool = {}
        for x in Member.objects.filter(group__code__in=self.select_type):
            merge_pool[x.code] = {
                "code": x.code,
                "name": x.full_name,
                "data": {},
                "transfer": {}
            }

        for k, v in total.items():
            if k in merge_pool:
                merge_pool[k]['data'] = v['data']

        for k, v in merge_pool.items():
            self.fill_row(count, v, row=current_row)
            count += 1
            current_row += 1

    @property
    def file_name(self):
        name = '{}_{}'.format(self.Meta.file_name, self.select_date.strftime('%Y_%b'))
        return '{}.xlsx'.format(name)
