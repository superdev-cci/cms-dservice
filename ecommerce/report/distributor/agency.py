from collections import OrderedDict
from datetime import date

from openpyxl.styles import PatternFill
from openpyxl.styles.numbers import FORMAT_NUMBER_COMMA_SEPARATED1, BUILTIN_FORMATS

from core.mixin import MonthMixIn
from core.report.excel import ExcelTemplateReport
from ecommerce.report.sale import AgFrBoughtSummaryReport, AgChildBoughtSummaryReport


class AgencyBoughtReport(ExcelTemplateReport, MonthMixIn):
    template_file = './templates/report/ecommerce/member_bought ag_fr.xlsx'
    margin = 20000
    fillRed = PatternFill("solid", fgColor='00f29a98')
    fillGreen = PatternFill("solid", fgColor='00b4f198')

    class Meta:
        title = 'Ag Qualified check'
        sheet_name = 'Summary'
        file_name = 'ag_check'
        head_file = 'Ag Qualified check'
        head_start_col = 4
        head_start_row = 6
        content_start_col = 1
        content_start_row = 7

    def __init__(self, *args, **kwargs):
        super(AgencyBoughtReport, self).__init__(**kwargs)
        self.select_date = kwargs.get('date', date.today())
        self.start, self.end = self.get_year_range(self.select_date)
        self.bough_history = AgFrBoughtSummaryReport(start=self.start, end=self.end, get_type='monthly')

    def create_head(self):
        self.work_sheet['C4'] = 'Agency'
        self.work_sheet['D5'] = date.today().strftime('%d/%b/%Y')
        start, end = self.get_year_range(self.select_date)
        current_col = self.Meta.head_start_col
        cell = self.get_cell(current_col, self.Meta.head_start_row)
        cell.border = self.style['full_border']
        for x in self._calculate_month_range(self.start, self.end):
            current = x.strftime('%Y-%b')
            cell.value = current
            cell = self.next_cell(cell, 1, 0)
        cell.value = 'Total'

    def build_row_meta(self, row_index, data, **kwargs):
        row_meta = [
            ('no', {'data': row_index, 'alignment': self.style['align_center']}),
            ('mcode', {'data': data['code']}),
            ('name', {'data': data['name']}),
        ]
        month_range = self._calculate_month_range(self.start, self.end)
        pass_count = 0
        for x in month_range:
            current_month = x.strftime('%Y-%b')
            pv_month = data['data'].get(current_month, 0)
            if pv_month >= self.margin:
                row_meta.append((current_month, {
                    'data': pv_month,
                    'number_format': BUILTIN_FORMATS[3],
                    'alignment': self.style['align_right'],
                    'fill': self.fillGreen
                }))
                pass_count += 1
            else:
                row_meta.append((current_month, {
                    'data': pv_month,
                    'number_format': BUILTIN_FORMATS[3],
                    'alignment': self.style['align_right'],
                    'fill': self.fillRed
                }))

        row_meta.append(('total', {
            'data': pass_count,
            'alignment': self.style['align_right'],
        }))
        return OrderedDict(row_meta)

    def process_data(self):
        count = 1
        current_row = self.Meta.content_start_row
        self.create_head()
        total = self.bough_history.total
        for k, v in total.items():
            self.fill_row(count, v, row=current_row)
            count += 1
            current_row += 1

    @property
    def file_name(self):
        name = '{}_{}'.format(self.Meta.file_name, self.select_date.strftime('%Y_%b'))
        return '{}.xlsx'.format(name)
